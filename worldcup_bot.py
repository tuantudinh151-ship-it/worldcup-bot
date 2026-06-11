import os, json, logging, asyncio, io
from datetime import datetime, timedelta
import pytz, requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import Application, CommandHandler, PollAnswerHandler, ContextTypes
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# ============================================================
#  ĐỌC CẤU HÌNH TỪ FILE .env (không cần sửa file này)
# ============================================================
def load_env():
    """Đọc config từ file .env (local) hoặc biến môi trường (Railway/server)"""
    env = {}
    # Đọc file .env nếu có (chạy local)
    if os.path.exists(".env"):
        with open(".env", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
    # Biến môi trường hệ thống (Railway, Render...) ghi đè lên .env
    for key in ["BOT_TOKEN", "GROUP_ID", "ADMIN_ID", "ODDS_API_KEY", "APIFOOTBALL_KEY", "GITHUB_TOKEN", "GITHUB_REPO", "TEST_MODE", "TEST_SPORT_KEY", "MISE_PYTHON_GITHUB_ATTESTATIONS"]:
        val = os.environ.get(key)
        if val:
            env[key] = val
    return env

_env = load_env()
BOT_TOKEN       = _env.get("BOT_TOKEN", "")
GROUP_ID        = int(_env.get("GROUP_ID", "0"))
ADMIN_ID        = int(_env.get("ADMIN_ID", "0"))
ODDS_API_KEY    = _env.get("ODDS_API_KEY", "")
APIFOOTBALL_KEY = _env.get("APIFOOTBALL_KEY", "")
GITHUB_TOKEN    = _env.get("GITHUB_TOKEN", "")
GITHUB_REPO     = _env.get("GITHUB_REPO", "")
TIMEZONE        = "Asia/Ho_Chi_Minh"
# TEST MODE: dùng trận giao hữu thay World Cup
# Đặt TEST_MODE=true trong Railway Variables để bật
TEST_MODE       = _env.get("TEST_MODE", "false").lower() == "true"
# Sport key có thể override qua biến môi trường TEST_SPORT_KEY
# Mặc định thử soccer_friendly_international, nếu không có thì dùng EPL
TEST_SPORT_KEY  = _env.get("TEST_SPORT_KEY", "soccer_friendly_international")

if not BOT_TOKEN or not GROUP_ID or not ADMIN_ID:
    print("LỖI: Thiếu thông tin cấu hình!")
    print("Local: chạy setup_config.bat để tạo file .env")
    print("Railway: vào Variables và thêm BOT_TOKEN, GROUP_ID, ADMIN_ID")
    exit(1)
# ============================================================

def build_rules() -> str:
    return (
        "<b>🏆 WORLD CUP 2026 — THỂ LỆ DỰ ĐOÁN 🏆</b>\n\n"
        "Chào mừng bạn tham gia nhóm dự đoán kết quả World Cup 2026!\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>📋 CÁCH THAM GIA</b>\n"
        "Gõ /thamgia trong nhóm để đăng ký (chỉ làm 1 lần)\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>⚽ CÁCH BÌNH CHỌN</b>\n"
        "Bot tự động gửi poll trước mỗi trận <b>20 tiếng</b>\n"
        "Poll được ghim lên đầu nhóm để dễ thấy\n\n"
        "Mỗi poll có 2 hoặc 3 lựa chọn:\n"
        "• <b>TRÊN</b> — đội chấp thắng kèo\n"
        "• <b>DƯỚI</b> — đội được chấp thắng kèo\n"
        "• <b>HÒA</b> — chỉ xuất hiện với kèo chẵn (0, 1, 2 trái)\n\n"
        "✅ Được đổi ý cho đến khi poll khóa\n"
        "🔒 Poll tự động khóa đúng giờ bóng lăn\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>🎯 KÈO CHÂU Á — CÁCH ĐỌC</b>\n"
        "Ví dụ: <b>Brazil chấp Argentina 0.5 trái</b>\n"
        "→ Brazil phải thắng ít nhất 1 trái mới thắng kèo\n"
        "→ Nếu hòa hoặc Argentina thắng → DƯỚI thắng kèo\n\n"
        "Ví dụ: <b>France chấp Germany 1 trái</b> (kèo chẵn)\n"
        "→ France thắng 2+ trái → TRÊN thắng\n"
        "→ France thắng đúng 1 trái → HÒA kèo\n"
        "→ Hòa hoặc Germany thắng → DƯỚI thắng\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>💸 PHÍ THUA CUỘC</b>\n"
        "• Vòng bảng: <b>50.000đ/trận</b>\n"
        "• Vòng loại trực tiếp: <b>100.000đ/trận</b>\n"
        "• Chung kết: <b>200.000đ/trận</b>\n\n"
        "⚠️ <b>Không bình chọn = tính thua</b>\n"
        "Tiền nợ được tổng hợp tự động, admin thông báo khi cần đóng\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>📊 LỆNH HỮU ÍCH</b>\n"
        "/lichthidau — Xem trận sắp tới\n"
        "/bangno — Bảng xếp hạng nợ\n"
        "/lichsu — Lịch sử bình chọn của bạn\n"
        "/thele — Xem lại thể lệ này\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "<b>🤖 KẾT QUẢ &amp; TÍNH TIỀN</b>\n"
        "Sau mỗi trận, bot tự động:\n"
        "✔️ Lấy kết quả từ nhà cái\n"
        "✔️ Tính thắng/thua theo kèo\n"
        "✔️ Thông báo vào nhóm\n"
        "✔️ Cập nhật bảng nợ\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n""<b>💳 THANH TOÁN TIỀN THUA</b>\n""Chuyển khoản về số tài khoản:\n""<b>0969984192</b> — Ngân hàng <b>Vietinbank</b>\n""Nội dung: <i>[Tên] đóng tiền WC2026</i>\n\n""Chúc mọi người may mắn! 🍀"
    )

RULES_TEXT = build_rules()
RULES_PARSE_MODE = "HTML"

FEE          = {"group": 50000, "knockout": 100000, "final": 200000}
# Sport key tự động theo mode
def get_sport_key():
    return TEST_SPORT_KEY if TEST_MODE else "soccer_fifa_world_cup"
POLL_BEFORE  = 20  # Gửi poll trước kickoff bao nhiêu tiếng
TZ           = pytz.timezone(TIMEZONE)
DATA_FILE    = "data.json"
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
#  THÔNG TIN VÒNG ĐẤU (World Cup 2026)
# ============================================================
def get_round_type(commence_time: datetime) -> str:
    """
    World Cup 2026: 12/06 - 20/07
    Test mode: tất cả trận giao hữu tính là vòng bảng (50k)
    """
    if TEST_MODE:
        return "group"  # Test mode: phí cố định 50k
    from datetime import date
    d = commence_time.astimezone(TZ).date()
    if d <= date(2026, 6, 27): return "group"
    if d <= date(2026, 7, 19): return "knockout"
    return "final"

# ============================================================
#  DATA
# ============================================================
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"matches": {}, "predictions": {}, "players": {}}



def is_admin(uid): return uid == ADMIN_ID

# ============================================================
#  SYNC DATA LÊN GITHUB
# ============================================================

def push_data_to_github():
    """Push data.json lên GitHub để backup, khôi phục khi Railway restart."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return
    try:
        import base64
        api = f"https://api.github.com/repos/{GITHUB_REPO}/contents/data.json"
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }
        # Đọc file hiện tại
        with open(DATA_FILE, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode()

        # Lấy SHA của file hiện tại trên GitHub (cần để update)
        r = requests.get(api, headers=headers, timeout=10)
        sha = r.json().get("sha") if r.status_code == 200 else None

        payload = {
            "message": f"auto: update data {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}",
            "content": content_b64
        }
        if sha:
            payload["sha"] = sha

        r = requests.put(api, headers=headers, json=payload, timeout=10)
        if r.status_code in [200, 201]:
            logger.info("Đã sync data.json lên GitHub")
        else:
            logger.error(f"Lỗi push GitHub: {r.status_code} {r.text[:100]}")
    except Exception as e:
        logger.error(f"push_data_to_github lỗi: {e}")


def pull_data_from_github():
    """Tải data.json từ GitHub về khi khởi động (khôi phục sau Railway restart)."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return
    try:
        import base64
        api = f"https://api.github.com/repos/{GITHUB_REPO}/contents/data.json"
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }
        r = requests.get(api, headers=headers, timeout=10)
        if r.status_code == 200:
            content_b64 = r.json().get("content", "")
            data_bytes = base64.b64decode(content_b64)
            with open(DATA_FILE, "wb") as f:
                f.write(data_bytes)
            logger.info("Đã khôi phục data.json từ GitHub")
        elif r.status_code == 404:
            logger.info("Chưa có data.json trên GitHub - bắt đầu mới")
        else:
            logger.error(f"Lỗi pull GitHub: {r.status_code}")
    except Exception as e:
        logger.error(f"pull_data_from_github lỗi: {e}")


def save_data(data):
    """Lưu data vào file và sync lên GitHub."""
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    push_data_to_github()

# ============================================================
#  LẤY KÈO TỪ THE ODDS API
# ============================================================
def fetch_available_sports():
    """Lấy danh sách sport keys có sẵn từ API."""
    try:
        r = requests.get(
            "https://api.the-odds-api.com/v4/sports",
            params={"apiKey": ODDS_API_KEY},
            timeout=10
        )
        if r.status_code == 200:
            return [s["key"] for s in r.json() if s.get("active")]
        return []
    except:
        return []

def fetch_odds(market="spreads"):
    sport_key = get_sport_key()
    try:
        r = requests.get(
            f"https://api.the-odds-api.com/v4/sports/{sport_key}/odds",
            params={"apiKey": ODDS_API_KEY, "regions": "eu",
                    "markets": market, "oddsFormat": "decimal", "dateFormat": "iso"},
            timeout=15
        )
        if r.status_code == 404 and TEST_MODE:
            # Sport key không tồn tại → thử tìm key giao hữu khác
            logger.warning(f"Sport key '{sport_key}' không có, tìm key thay thế...")
            available = fetch_available_sports()
            fallbacks = [k for k in available if "friendly" in k or "international" in k]
            if not fallbacks:
                # Không có giao hữu, thử giải đang có trận
                fallbacks = [k for k in available if "soccer" in k]
            if fallbacks:
                logger.info(f"Dùng sport key thay thế: {fallbacks[0]}")
                r = requests.get(
                    f"https://api.the-odds-api.com/v4/sports/{fallbacks[0]}/odds",
                    params={"apiKey": ODDS_API_KEY, "regions": "eu",
                            "markets": market, "oddsFormat": "decimal", "dateFormat": "iso"},
                    timeout=15
                )
            else:
                logger.error("Không tìm được sport key thay thế")
                return []
        if r.status_code != 200:
            logger.error(f"Odds API lỗi: {r.status_code} {r.text}")
            return []
        return r.json()
    except Exception as e:
        logger.error(f"fetch_odds lỗi: {e}")
        return []

def parse_asian_handicap(game: dict) -> dict | None:
    try:
        priority = ["pinnacle", "betfair_ex_eu", "betonlineag", "mybookieag"]
        chosen_bm = None
        for bm_key in priority:
            for bm in game.get("bookmakers", []):
                if bm["key"] == bm_key:
                    chosen_bm = bm
                    break
            if chosen_bm: break
        if not chosen_bm and game.get("bookmakers"):
            chosen_bm = game["bookmakers"][0]
        if not chosen_bm: return None

        market = next((m for m in chosen_bm["markets"] if m["key"] == "spreads"), None)
        if not market: return None
        outcomes = market["outcomes"]
        if len(outcomes) < 2: return None

        home = next((o for o in outcomes if o["name"] == game["home_team"]), outcomes[0])
        away = next((o for o in outcomes if o["name"] == game["away_team"]), outcomes[1])

        return {
            "home_team":     game["home_team"],
            "away_team":     game["away_team"],
            "commence_time": game["commence_time"],
            "game_id":       game["id"],  # UUID gốc để tra cứu
            "bookmaker":     chosen_bm["title"],
            "handicap":      home.get("point", 0),
            "home_odds":     home["price"],
            "away_odds":     away["price"],
        }
    except Exception as e:
        logger.error(f"parse lỗi: {e}")
        return None

def normalize_handicap(handicap: float) -> float:
    """
    Làm tròn kèo:
    0.25 → 0.5 | 0.75 → 0.5 | 1.25 → 1.5 | 1.75 → 1.5 | v.v.
    Giữ nguyên: 0, 0.5, 1, 1.5, 2...
    """
    sign  = -1 if handicap < 0 else 1
    abs_h = abs(handicap)
    frac  = round(abs_h % 1, 4)
    if frac in (0.25, 0.75):
        abs_h = int(abs_h) + 0.5
    return round(sign * abs_h, 2)

def has_draw_option(handicap: float) -> bool:
    """Kèo nguyên (0, 1, 2...) → có lựa chọn HÒA"""
    h = normalize_handicap(handicap)
    return h % 1 == 0

def format_handicap(handicap: float, home: str, away: str) -> str:
    h = normalize_handicap(handicap)
    if h == 0:        return "Kèo chẵn 0 (có thêm HÒA)"
    elif h < 0:       return f"{home} chấp {abs(h)} trái"
    else:             return f"{away} chấp {h} trái"

def _truncate_opt(s: str) -> str:
    """Cắt option về tối đa 100 ký tự (giới hạn Telegram)."""
    return s if len(s) <= 100 else s[:97] + "..."

def build_poll_options(m: dict) -> list[str]:
    """
    Kèo nguyên (0, 1, 2...): 3 lựa chọn TRÊN / HÒA / DƯỚI
    Kèo lẻ (0.5, 1.5...): 2 lựa chọn TRÊN / DƯỚI
    Kèo 0.25/0.75 đã làm tròn → 0.5 → 2 lựa chọn
    Sai = mất 100% trong mọi trường hợp.
    """
    h    = normalize_handicap(m["handicap"])
    home = m["home_team"]
    away = m["away_team"]
    ho   = m["home_odds"]
    ao   = m["away_odds"]

    if h % 1 == 0:
        if h == 0:
            return [
                f"TRÊN - {home} thắng ({ho})",
                f"HÒA - Trận đấu hòa",
                f"DƯỚI - {away} thắng ({ao})",
            ]
        else:
            return [
                f"TRÊN - {home} ({ho})",
                f"HÒA KÈO - Trận hòa đúng kèo",
                f"DƯỚI - {away} ({ao})",
            ]
    else:
        return [
            f"TRÊN - {home} ({ho})",
            f"DƯỚI - {away} ({ao})",
        ]

def get_choice_from_index(index: int, m: dict) -> str:
    """Chuyển index bình chọn → home/draw/away"""
    if has_draw_option(m.get("handicap", 0)):
        return ["home", "draw", "away"][index] if index < 3 else "away"
    else:
        return "home" if index == 0 else "away"

# ============================================================
#  GỬI VÀ KHÓA POLL
# ============================================================

async def remind_unvoted_job(match_id: str, app):
    """Nhắc những ai chưa bình chọn, chạy trước kickoff 1 tiếng."""
    data = load_data()
    if match_id not in data["matches"]:
        return
    m = data["matches"][match_id]
    if m.get("locked") or m.get("result"):
        return

    members = data.get("members", {})
    preds   = data["predictions"].get(match_id, {})
    unvoted = [member["name"] for uid, member in members.items() if uid not in preds]

    if not unvoted:
        return  # Mọi người đã bình chọn

    kickoff = datetime.fromisoformat(m["kickoff"])
    fee     = FEE.get(m["round"], 50000)
    mentions = ", ".join(unvoted)

    msg = (
        f"⏰ NHẮC NHỞ - Còn 1 tiếng nữa bóng lăn!\n"
        f"Trận {match_id}: {m['home_team']} vs {m['away_team']}\n"
        f"Kickoff: {kickoff.strftime('%H:%M %d/%m')}\n\n"
        f"CHƯA BÌNH CHỌN ({len(unvoted)} người):\n{mentions}\n\n"
        f"⚠️ Không bình chọn sẽ bị tính THUA -{fee:,}đ!\n"
        f"Bình chọn ngay tại poll đã ghim phía trên."
    )
    try:
        await app.bot.send_message(chat_id=GROUP_ID, text=msg)
        logger.info(f"Đã nhắc {len(unvoted)} người chưa bình chọn trận {match_id}")
    except Exception as e:
        logger.error(f"Lỗi nhắc nhở {match_id}: {e}")



async def auto_send_poll_job(match_id: str, app):
    """Gửi poll vào nhóm và ghim lên đầu"""
    data = load_data()
    if match_id not in data["matches"]:
        return
    m = data["matches"][match_id]
    if m.get("poll_message_id"):
        return  # Đã gửi rồi

    kickoff = datetime.fromisoformat(m["kickoff"])
    fee = FEE.get(m["round"], 50000)
    hcap_text = format_handicap(m["handicap"], m["home_team"], m["away_team"])
    options = build_poll_options(m)
    even_note = "\nLưu ý: Kèo chẵn - có lựa chọn HÒA!" if has_draw_option(m["handicap"]) else ""

    question = (
        f"BÌNH CHỌN {match_id}: {m['home_team']} vs {m['away_team']}\n"
        f"Kickoff: {kickoff.strftime('%d/%m/%Y %H:%M')} | Phí thua: {fee:,}đ\n"
        f"Kèo: {hcap_text}{even_note}"
    )
    # Telegram giới hạn question 300 ký tự
    if len(question) > 300:
        question = question[:297] + "..."

    try:
        # Poll ẨN DANH: thành viên không thấy ai vote gì, không thấy số phiếu
        # cho đến khi poll đóng (giới hạn của Telegram với anonymous poll)
        msg = await app.bot.send_poll(
            chat_id=GROUP_ID,
            question=question,
            options=options,
            is_anonymous=False,            # Công khai - thấy ai bình chọn gì
            allows_multiple_answers=False,
            open_period=None               # Không tự đóng, bot khóa đúng kickoff
        )

        # Ghim poll lên đầu nhóm
        await app.bot.pin_chat_message(
            chat_id=GROUP_ID,
            message_id=msg.message_id,
            disable_notification=True   # Ghim không thông báo ồn ào
        )

        data["matches"][match_id]["poll_id"] = msg.poll.id
        data["matches"][match_id]["poll_message_id"] = msg.message_id
        data["matches"][match_id]["has_draw_option"] = has_draw_option(m["handicap"])
        save_data(data)
        logger.info(f"Đã gửi + ghim poll trận {match_id}")
    except Exception as e:
        logger.error(f"Lỗi gửi poll {match_id}: {e}")


async def lock_poll_job(match_id: str, app):
    """Khóa poll + bỏ ghim khi kickoff"""
    data = load_data()
    if match_id not in data["matches"]:
        return
    m = data["matches"][match_id]
    if m["locked"] or not m.get("poll_message_id"):
        return
    try:
        # Khóa poll
        await app.bot.stop_poll(chat_id=GROUP_ID, message_id=m["poll_message_id"])

        # Bỏ ghim
        await app.bot.unpin_chat_message(
            chat_id=GROUP_ID,
            message_id=m["poll_message_id"]
        )

        data["matches"][match_id]["locked"] = True
        save_data(data)

        await app.bot.send_message(
            chat_id=GROUP_ID,
            text=(
                f"KHÓA BÌNH CHỌN - Trận {match_id}\n"
                f"{m['home_team']} vs {m['away_team']}\n"
                f"Bóng lăn rồi! Chờ kết quả..."
            )
        )
        logger.info(f"Đã khóa + bỏ ghim poll trận {match_id}")
    except Exception as e:
        logger.error(f"Lỗi khóa poll {match_id}: {e}")


# ============================================================
#  LỆNH ADMIN
# ============================================================

async def cmd_newmatch(update, context):
    if not is_admin(update.effective_user.id): return
    args = context.args
    if len(args) < 9:
        await update.message.reply_text(
            'Cu phap:\n'
            '/themtran <id> <doi1> vs <doi2> <ngay> <gio> <vong> <handicap> <odds_tren> <odds_duoi>\n'
            'Vi du:\n'
            '/themtran TEST Brazil vs Argentina 2026-06-07 22:05 group -0.5 1.95 1.95\n'
            'handicap: -0.5=nha chap nua | 0=keo chan | 1=khach chap 1'
        )
        return
    try:
        match_id   = args[0].upper()
        vs_idx     = args.index('vs')
        team1      = ' '.join(args[1:vs_idx])
        team2      = ' '.join(args[vs_idx+1:-6])
        date_str   = args[-6]
        time_str   = args[-5]
        round_type = args[-4]
        handicap   = float(args[-3])
        home_odds  = float(args[-2])
        away_odds  = float(args[-1])
        if round_type not in ['group', 'knockout', 'final']:
            raise ValueError('Vong phai la: group / knockout / final')
        kickoff = TZ.localize(datetime.strptime(date_str + ' ' + time_str, '%Y-%m-%d %H:%M'))
        data = load_data()
        if match_id in data['matches']:
            await update.message.reply_text('Tran ' + match_id + ' da ton tai.')
            return
        h_norm = normalize_handicap(handicap)
        data['matches'][match_id] = {
            'home_team': team1, 'away_team': team2,
            'kickoff': kickoff.isoformat(), 'round': round_type,
            'handicap': h_norm, 'home_odds': home_odds, 'away_odds': away_odds,
            'bookmaker': 'Thu cong', 'has_draw_option': has_draw_option(h_norm),
            'result': None, 'poll_id': None, 'poll_message_id': None, 'locked': False
        }
        save_data(data)
        scheduler = context.bot_data.get('scheduler')
        now = datetime.now(TZ)
        send_time = kickoff - timedelta(hours=POLL_BEFORE)
        if send_time > now and scheduler:
            scheduler.add_job(auto_send_poll_job, 'date', run_date=send_time,
                args=[match_id, context.application], id='send_' + match_id, replace_existing=True)
        if kickoff > now and scheduler:
            scheduler.add_job(lock_poll_job, 'date', run_date=kickoff,
                args=[match_id, context.application], id='lock_' + match_id, replace_existing=True)
        fee = FEE.get(round_type, 50000)
        hcap_txt = format_handicap(h_norm, team1, team2)
        await update.message.reply_text(
            'Da tao tran ' + match_id + ':\n' +
            team1 + ' vs ' + team2 + '\n' +
            'Kickoff: ' + kickoff.strftime('%d/%m/%Y %H:%M') + '\n' +
            'Keo: ' + hcap_txt + ' | Phi thua: ' + str(fee) + 'd\n' +
            'Go /guibinhchon ' + match_id + ' de gui poll vao nhom.'
        )
    except Exception as e:
        await update.message.reply_text('Loi: ' + str(e))


async def cmd_fetchodds(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem kèo mới từ API. /xemkeo"""
    if not is_admin(update.effective_user.id): return
    await update.message.reply_text("Đang lấy kèo...")

    games = fetch_odds("spreads")
    if not games:
        await update.message.reply_text("Không lấy được dữ liệu. Kiểm tra API key.")
        return

    now = datetime.now(TZ)
    upcoming = []
    for g in games:
        p = parse_asian_handicap(g)
        if not p: continue
        kickoff = datetime.fromisoformat(p["commence_time"].replace("Z", "+00:00")).astimezone(TZ)
        if kickoff > now:
            upcoming.append((kickoff, p))
    upcoming.sort(key=lambda x: x[0])

    if not upcoming:
        await update.message.reply_text("Không có trận nào sắp tới.")
        return

    msg = f"DANH SÁCH TRẬN ({len(upcoming)} trận)\n/laykeo để tự động tạo tất cả\n" + "="*32 + "\n"
    for i, (kickoff, p) in enumerate(upcoming[:12], 1):
        hcap = format_handicap(p["handicap"], p["home_team"], p["away_team"])
        even = " [CHẴN+HÒA]" if has_draw_option(p["handicap"]) else ""
        msg += (
            f"{i}. {p['home_team']} vs {p['away_team']}\n"
            f"   {kickoff.strftime('%d/%m %H:%M')} | {hcap}{even}\n"
            f"   Tỷ lệ: {p['home_odds']} / {p['away_odds']} | {p['bookmaker']}\n\n"
        )
    await update.message.reply_text(msg)


async def cmd_autosetup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tự động tạo trận + lên lịch gửi poll và khóa. /laykeo [số_trận]"""
    if not is_admin(update.effective_user.id): return
    limit = int(context.args[0]) if context.args else 999
    await update.message.reply_text(f"Đang lấy kèo và tạo lịch...")

    games = fetch_odds("spreads")
    if not games:
        await update.message.reply_text("Không lấy được dữ liệu API.")
        return

    data = load_data()
    now = datetime.now(TZ)
    scheduler = context.bot_data.get("scheduler")
    created = skipped = 0

    for g in sorted(games, key=lambda x: x["commence_time"]):
        if created >= limit: break
        p = parse_asian_handicap(g)
        if not p: continue

        kickoff = datetime.fromisoformat(p["commence_time"].replace("Z", "+00:00")).astimezone(TZ)
        if kickoff <= now: continue

        # Tạo ID dạng WC001, WC002... theo thứ tự kickoff
        prefix = "WC" if not TEST_MODE else "TS"
        # Tính base một lần, tăng theo created trong vòng lặp
        if created == 0:
            existing_nums = [
                int(mid[len(prefix):]) for mid in data["matches"]
                if mid.startswith(prefix) and mid[len(prefix):].isdigit()
            ]
            context.bot_data["_next_match_num"] = max(existing_nums, default=0)
        base_num = context.bot_data.get("_next_match_num", 0)
        match_id = f"{prefix}{base_num + created + 1:03d}"

        # Kiểm tra trùng trận (cùng đội + cùng giờ)
        duplicate = any(
            m["home_team"] == p["home_team"] and
            m["away_team"] == p["away_team"] and
            m["kickoff"][:10] == kickoff.isoformat()[:10]
            for m in data["matches"].values()
        )
        if duplicate:
            skipped += 1
            continue

        round_type = get_round_type(kickoff)
        data["matches"][match_id] = {
            "home_team":      p["home_team"],
            "away_team":      p["away_team"],
            "kickoff":        kickoff.isoformat(),
            "round":          round_type,
            "handicap":       p["handicap"],
            "home_odds":      p["home_odds"],
            "away_odds":      p["away_odds"],
            "bookmaker":      p["bookmaker"],
            "has_draw_option": has_draw_option(p["handicap"]),
            "result":         None,
            "poll_id":        None,
            "poll_message_id": None,
            "locked":         False
        }

        # Lên lịch gửi poll trước POLL_BEFORE tiếng
        send_time = kickoff - timedelta(hours=POLL_BEFORE)
        if send_time > now and scheduler:
            scheduler.add_job(auto_send_poll_job, "date", run_date=send_time,
                args=[match_id, context.application],
                id=f"send_{match_id}", replace_existing=True)

        # Lên lịch khóa + bỏ ghim lúc kickoff
        if scheduler:
            scheduler.add_job(lock_poll_job, "date", run_date=kickoff,
                args=[match_id, context.application],
                id=f"lock_{match_id}", replace_existing=True)

        context.bot_data["_next_match_num"] = base_num + created + 1
        created += 1

    save_data(data)
    await update.message.reply_text(
        f"Đã tạo {created} trận, bỏ qua {skipped} trận đã có.\n"
        f"Poll tự gửi trước {POLL_BEFORE} tiếng, ghim lên đầu nhóm.\n"
        f"Poll tự khóa + bỏ ghim đúng giờ kickoff.\n"
        f"Dùng /lichthidau để xem danh sách."
    )


async def cmd_sendpoll(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Gửi poll thủ công. /guibinhchon <mã_trận>"""
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /guibinhchon <mã_trận>")
        return
    match_id = context.args[0].upper()
    await auto_send_poll_job(match_id, context.application)
    await update.message.reply_text(f"Đã gửi + ghim poll trận {match_id}.")


async def cmd_lockpoll(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Khóa poll thủ công. /khoabinhchon <mã_trận>"""
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /khoabinhchon <mã_trận>")
        return
    match_id = context.args[0].upper()
    await lock_poll_job(match_id, context.application)
    await update.message.reply_text(f"Đã khóa + bỏ ghim poll trận {match_id}.")


async def cmd_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Nhập kết quả kèo. /ketqua <mã_trận> <home|draw|away>
    home  = đội nhà thắng kèo (TRÊN)
    draw  = hòa kèo (chỉ với kèo chẵn, hoàn 50%)
    away  = đội khách thắng kèo (DƯỚI)
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text("Cú pháp: /ketqua <mã_trận> <home|draw|away>")
        return

    match_id = context.args[0].upper()
    result   = context.args[1].lower()
    if result not in ["home", "draw", "away"]:
        await update.message.reply_text("Kết quả phải là: home / draw / away")
        return

    data = load_data()
    if match_id not in data["matches"]:
        await update.message.reply_text(f"Không tìm thấy trận {match_id}.")
        return

    m = data["matches"][match_id]
    data["matches"][match_id]["result"] = result
    fee  = FEE.get(m["round"], 50000)

    winners, losers = [], []
    for user_id, pred in data["predictions"].get(match_id, {}).items():
        player = data["players"].setdefault(user_id, {"name": pred["name"], "debt": 0})
        player["name"] = pred["name"]
        choice = pred["choice"]

        if choice == result:
            winners.append(pred["name"])
        else:
            # Sai = mất 100% trong mọi trường hợp
            losers.append(pred["name"])
            player["debt"] = player.get("debt", 0) + fee

    save_data(data)

    result_label = {"home": m["home_team"], "draw": "Hòa kèo", "away": m["away_team"]}[result]
    hcap_text = format_handicap(m["handicap"], m["home_team"], m["away_team"])

    msg = (
        f"KẾT QUẢ TRẬN {match_id}\n"
        f"{m['home_team']} vs {m['away_team']}\n"
        f"Kèo: {hcap_text}\n"
        f"Kết quả kèo: {result_label}\n\n"
    )
    if winners: msg += f"THẮNG ({len(winners)}): {', '.join(winners)}\n"
    if losers:  msg += f"THUA -{fee:,}đ ({len(losers)}): {', '.join(losers)}\n"
    if not data["predictions"].get(match_id):
        msg += "Không có ai bình chọn trận này."

    await context.bot.send_message(chat_id=GROUP_ID, text=msg)
    await update.message.reply_text("Đã cập nhật kết quả.")


# ============================================================
#  LỆNH CHO TẤT CẢ
# ============================================================

async def cmd_standings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = load_data()
    if not data["players"]:
        await update.message.reply_text("Chưa có dữ liệu.")
        return
    sorted_p = sorted(data["players"].items(), key=lambda x: x[1].get("debt", 0), reverse=True)
    msg = "BẢNG XẾP HẠNG NỢ\n" + "="*28 + "\n"
    for i, (uid, p) in enumerate(sorted_p, 1):
        debt = p.get("debt", 0)
        icon = "🔴" if debt > 0 else "🟢"
        msg += f"{i}. {icon} {p['name']}: {'-' if debt>0 else ''}{debt:,}đ\n"
    await update.message.reply_text(msg)


async def cmd_matches(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = load_data()
    now = datetime.now(TZ)
    upcoming = [
        (mid, m, datetime.fromisoformat(m["kickoff"]))
        for mid, m in data["matches"].items()
        if datetime.fromisoformat(m["kickoff"]) > now and not m.get("result")
    ]
    upcoming.sort(key=lambda x: x[2])
    if not upcoming:
        await update.message.reply_text("Không có trận nào sắp tới.")
        return
    msg = f"CÁC TRẬN SẮP TỚI ({len(upcoming)} trận)\n" + "="*28 + "\n"
    for mid, m, kickoff in upcoming[:10]:
        status = "📌 Đã ghim" if m.get("poll_message_id") and not m["locked"] else ("🔒 Đã khóa" if m["locked"] else "⏳ Chờ gửi")
        hcap = format_handicap(m.get("handicap", 0), m["home_team"], m["away_team"])
        even = " [+HÒA]" if m.get("has_draw_option") else ""
        msg += f"{mid}: {m['home_team']} vs {m['away_team']}\n"
        msg += f"   {kickoff.strftime('%d/%m %H:%M')} | {status} | {hcap}{even}\n"
    await update.message.reply_text(msg)


async def cmd_mypredictions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    data = load_data()
    history = []
    for mid, preds in data["predictions"].items():
        if user_id in preds:
            m = data["matches"].get(mid, {})
            result = m.get("result")
            pred   = preds[user_id]["choice"]
            label  = {"home": m.get("home_team","?"), "draw": "Hòa", "away": m.get("away_team","?")}.get(pred, pred)
            outcome = ("THẮNG" if pred == result else "THUA") if result else "Chờ kết quả"
            history.append(f"{mid}: Chọn {label} → {outcome}")
    if not history:
        await update.message.reply_text("Bạn chưa bình chọn trận nào.")
        return
    debt = data["players"].get(user_id, {}).get("debt", 0)
    msg = "LỊCH SỬ BÌNH CHỌN\n" + "\n".join(history) + f"\n\nTổng nợ: {debt:,}đ"
    await update.message.reply_text(msg)


async def cmd_bieudo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Vẽ biểu đồ diễn biến nợ theo thời gian. /bieudo"""
    if not is_admin(update.effective_user.id): return
    data = load_data()
    members = data.get("members", {})
    if not members:
        await update.message.reply_text("Chưa có thành viên nào.")
        return

    await update.message.reply_text("Đang vẽ biểu đồ...")

    # Lấy trận theo thứ tự thời gian
    done = [(mid, m) for mid, m in data["matches"].items() if m.get("result")]
    done.sort(key=lambda x: x[1]["kickoff"])

    if not done:
        await update.message.reply_text("Chưa có trận nào có kết quả để vẽ.")
        return

    # Tính nợ tích lũy của từng người qua từng trận
    cumulative = {uid: [0] for uid in members}
    labels = ["Bắt đầu"]

    for mid, m in done:
        result = m["result"]
        fee    = FEE.get(m["round"], 50000)
        preds  = data["predictions"].get(mid, {})
        labels.append(mid)
        for uid in members:
            prev = cumulative[uid][-1]
            if uid in preds:
                add = 0 if preds[uid]["choice"] == result else fee
            else:
                add = fee  # không bình chọn = thua
            cumulative[uid].append(prev + add)

    # Vẽ biểu đồ
    try:
        plt.figure(figsize=(12, 7))
        x = range(len(labels))
        for uid, member in members.items():
            plt.plot(x, cumulative[uid], marker="o", linewidth=2, label=member["name"])

        plt.xlabel("Trận đấu")
        plt.ylabel("Nợ tích lũy (VNĐ)")
        plt.title("DIỄN BIẾN NỢ THEO TRẬN - WORLD CUP 2026")
        plt.xticks(x, labels, rotation=45, ha="right")
        plt.legend(loc="upper left", fontsize=9)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()

        # Format trục y theo nghìn đồng
        ax = plt.gca()
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, p: f"{int(v):,}"))

        chart_path = "debt_chart.png"
        plt.savefig(chart_path, dpi=100, bbox_inches="tight")
        plt.close()

        # Gửi vào nhóm và cho admin
        with open(chart_path, "rb") as f:
            await context.bot.send_photo(
                chat_id=GROUP_ID,
                photo=f,
                caption=f"📈 Diễn biến nợ qua {len(done)} trận đã đấu"
            )
        await update.message.reply_text("Đã gửi biểu đồ vào nhóm!")
    except Exception as e:
        await update.message.reply_text(f"Lỗi vẽ biểu đồ: {e}")



async def cmd_thongke(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bảng xếp hạng VUA DỰ ĐOÁN - ai đoán đúng nhiều nhất. /thongke"""
    data = load_data()
    members = data.get("members", {})
    if not members:
        await update.message.reply_text("Chưa có thành viên nào.")
        return

    # Tính thống kê từng người
    stats = {}
    for uid, member in members.items():
        stats[uid] = {"name": member["name"], "win": 0, "lose": 0, "novote": 0, "streak": 0, "max_streak": 0}

    # Duyệt các trận đã có kết quả theo thứ tự thời gian
    done = [(mid, m) for mid, m in data["matches"].items() if m.get("result")]
    done.sort(key=lambda x: x[1]["kickoff"])

    for mid, m in done:
        result = m["result"]
        preds  = data["predictions"].get(mid, {})
        for uid in members:
            if uid not in stats:
                continue
            if uid in preds:
                if preds[uid]["choice"] == result:
                    stats[uid]["win"] += 1
                    stats[uid]["streak"] += 1
                    stats[uid]["max_streak"] = max(stats[uid]["max_streak"], stats[uid]["streak"])
                else:
                    stats[uid]["lose"] += 1
                    stats[uid]["streak"] = 0
            else:
                stats[uid]["novote"] += 1
                stats[uid]["streak"] = 0

    # Sắp xếp theo số trận thắng, rồi tỷ lệ thắng
    def win_rate(s):
        total = s["win"] + s["lose"] + s["novote"]
        return s["win"] / total if total > 0 else 0

    ranked = sorted(stats.values(),
                    key=lambda s: (s["win"], win_rate(s)),
                    reverse=True)

    medals = ["🥇", "🥈", "🥉"]
    msg = "🏆 BẢNG XẾP HẠNG VUA DỰ ĐOÁN\n" + "="*30 + "\n\n"
    for i, s in enumerate(ranked):
        total = s["win"] + s["lose"] + s["novote"]
        rate  = round(win_rate(s) * 100)
        icon  = medals[i] if i < 3 else f"{i+1}."
        msg += (
            f"{icon} {s['name']}\n"
            f"    Thắng: {s['win']} | Thua: {s['lose']} | KBQ: {s['novote']}\n"
            f"    Tỷ lệ đúng: {rate}% | Chuỗi thắng max: {s['max_streak']}\n\n"
        )

    if len(done) == 0:
        msg += "Chưa có trận nào có kết quả."
    else:
        msg += f"Tổng số trận đã đấu: {len(done)}"

    await update.message.reply_text(msg)


async def cmd_mystat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Thống kê cá nhân chi tiết. /thongketoi"""
    uid  = str(update.effective_user.id)
    data = load_data()

    win = lose = novote = streak = max_streak = 0
    done = [(mid, m) for mid, m in data["matches"].items() if m.get("result")]
    done.sort(key=lambda x: x[1]["kickoff"])

    for mid, m in done:
        result = m["result"]
        preds  = data["predictions"].get(mid, {})
        if uid in preds:
            if preds[uid]["choice"] == result:
                win += 1; streak += 1; max_streak = max(max_streak, streak)
            else:
                lose += 1; streak = 0
        elif uid in data.get("members", {}):
            novote += 1; streak = 0

    total = win + lose + novote
    if total == 0:
        await update.message.reply_text("Bạn chưa tham gia trận nào có kết quả.")
        return

    rate = round(win / total * 100)
    debt = data["players"].get(uid, {}).get("debt", 0)
    paid = data.get("members", {}).get(uid, {}).get("paid", 0)

    # Thanh tiến trình tỷ lệ thắng
    filled = round(rate / 10)
    bar = "█" * filled + "░" * (10 - filled)

    msg = (
        f"📊 THỐNG KÊ CỦA BẠN\n" + "="*28 + "\n\n"
        f"Tổng trận: {total}\n"
        f"✅ Thắng: {win}\n"
        f"❌ Thua: {lose}\n"
        f"⚠️ Không bình chọn: {novote}\n\n"
        f"Tỷ lệ đúng: {rate}%\n"
        f"{bar}\n\n"
        f"🔥 Chuỗi thắng dài nhất: {max_streak}\n"
        f"💰 Tổng nợ: {debt:,}đ | Đã đóng: {paid:,}đ\n"
        f"Còn lại: {max(0, debt-paid):,}đ"
    )
    await update.message.reply_text(msg)



async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    is_ad = is_admin(update.effective_user.id)
    msg = (
        "HƯỚNG DẪN BOT WORLD CUP 2026\n" + "="*30 + "\n\n"
        "LỆNH THÀNH VIÊN:\n"
        "/thamgia — Đăng ký tham gia\n"
        "/lichthidau — Xem trận sắp tới\n"
        "/bangno — Bảng xếp hạng nợ\n"
        "/lichsu — Lịch sử bình chọn của bạn\n"
        "/thele — Xem thể lệ chơi\n\n"
    )
    if is_ad:
        msg += (
            "LỆNH ADMIN — VẬN HÀNH:\n"
            "/laykeo [số] — Lấy lịch + kèo tự động\n"
            "/xemkeo — Xem kèo từ nhà cái\n"
            "/capnhatkeo — Cập nhật kèo mới nhất\n"
            "/themtran — Tạo trận thủ công\n"
            "/guibinhchon <id> — Gửi + ghim poll\n"
            "/khoabinhchon <id> — Khóa poll\n"
            "/laykequa — Lấy kết quả tự động\n"
            "/ketqua <id> <home|draw|away> — Nhập kết quả tay\n"
            "/xuatfile — Xuất Excel tính tiền\n\n"
            "LỆNH ADMIN — THÀNH VIÊN & TIỀN:\n"
            "/danhsach — Danh sách thành viên\n"
            "/themthanhvien <id> <tên> — Thêm thủ công\n"
            "/xoathanhvien <id> — Xóa thành viên\n"
            "/dathanhtoan <tên> <tiền> — Xác nhận đóng\n"
            "/conno — Danh sách còn nợ\n"
            "/luudata — Backup lên GitHub\n\n"
            "LỆNH ADMIN — SỬA DỮ LIỆU:\n"
            "/xemtran <id> — Chi tiết trận\n"
            "/suatran <id> <field> <giá trị> — Sửa trận\n"
            "/xoatran <id> — Xóa trận\n"
            "/suakequa <id> <kết quả> — Sửa kết quả\n"
            "/suano <tên> <tiền> — Sửa nợ\n"
            "/suadathanhtoan <tên> <tiền> — Sửa tiền đã đóng\n\n"
            "LỆNH ADMIN — HỆ THỐNG:\n"
            "/chedo — Xem chế độ test/WC\n"
            "/guithele — Gửi thể lệ vào nhóm\n"
            "/xoatrancu <TS|WC|all> — Xóa trận theo loại\n"
            "/danhsachgiai — Danh sách giải có kèo\n"
            "/resetwc confirm — Reset sang World Cup\n"
            "/xoahetdata confirm — Xóa toàn bộ\n\n"
        )
    msg += (
        "PHÍ THUA:\n"
        "Vòng bảng: 50,000đ | Loại trực tiếp: 100,000đ | Chung kết: 200,000đ\n\n"
        "KÈO CHÂU Á:\n"
        "TRÊN = đội chấp thắng kèo\n"
        "DƯỚI = đội được chấp thắng kèo\n"
        "HÒA = chỉ ở kèo nguyên (0, 1, 2...)\n"
        "Sai = mất 100%, không bình chọn = tính thua"
    )
    await update.message.reply_text(msg)


# ============================================================
#  XỬ LÝ BÌNH CHỌN (ẨN DANH)
# ============================================================

async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Lưu bình chọn ẨN DANH - chỉ bot nhận được thông tin, nhóm không biết ai chọn gì.
    Người dùng CHỈ ĐƯỢC CHỌN 1 LẦN - không đổi ý được.
    Số phiếu bầu ẩn cho đến khi poll đóng (tính năng mặc định của Telegram anonymous poll).
    """
    answer  = update.poll_answer
    poll_id = answer.poll_id
    user    = answer.user
    user_id = str(user.id)

    # Người dùng bỏ vote -> xóa bình chọn cũ
    if not answer.option_ids:
        data = load_data()
        match_id = next((mid for mid, m in data["matches"].items() if m.get("poll_id") == poll_id), None)
        if match_id:
            data["predictions"].get(match_id, {}).pop(user_id, None)
            save_data(data)
        return

    option_index = answer.option_ids[0]
    data = load_data()
    match_id = next((mid for mid, m in data["matches"].items() if m.get("poll_id") == poll_id), None)
    if not match_id or data["matches"][match_id]["locked"]:
        return

    m = data["matches"][match_id]

    choice = get_choice_from_index(option_index, m)
    # Cho phép đổi ý - ghi đè bình chọn cũ
    data["predictions"].setdefault(match_id, {})[user_id] = {
        "name": user.full_name, "choice": choice
    }
    data["players"].setdefault(user_id, {"name": user.full_name, "debt": 0})["name"] = user.full_name
    save_data(data)
    logger.info(f"{user.full_name} bình chọn '{choice}' cho trận {match_id}")



async def cmd_editresult(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Sửa kết quả trận đã có, tính lại nợ chính xác.
    /suakequa <match_id> <home|draw|away>
    Bot sẽ hoàn lại nợ cũ rồi tính lại từ đầu.
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Cú pháp: /suakequa <match_id> <home|draw|away>\n"
            "Ví dụ: /suakequa TEST01 home\n\n"
            "Bot sẽ hoàn lại nợ cũ và tính lại từ đầu."
        )
        return

    match_id   = context.args[0].upper()
    new_result = context.args[1].lower()
    if new_result not in ["home", "draw", "away"]:
        await update.message.reply_text("Kết quả phải là: home / draw / away")
        return

    data = load_data()
    if match_id not in data["matches"]:
        await update.message.reply_text(f"Không tìm thấy trận {match_id}.")
        return

    m          = data["matches"][match_id]
    old_result = m.get("result")
    fee        = FEE.get(m["round"], 50000)
    members    = data.get("members", {})
    preds      = data["predictions"].get(match_id, {})

    # Bước 1: Hoàn lại nợ cũ
    if old_result:
        for uid, member in members.items():
            player = data["players"].get(uid)
            if not player:
                continue
            if uid in preds:
                if preds[uid]["choice"] != old_result:
                    # Người này đã bị tính thua cũ → hoàn lại
                    player["debt"] = max(0, player.get("debt", 0) - fee)
            else:
                # Không bình chọn → đã bị tính thua cũ → hoàn lại
                player["debt"] = max(0, player.get("debt", 0) - fee)

    # Bước 2: Tính lại với kết quả mới
    winners, losers, no_vote = [], [], []
    for uid, member in members.items():
        player = data["players"].setdefault(uid, {"name": member["name"], "debt": 0})
        player["name"] = member["name"]
        if uid in preds:
            choice = preds[uid]["choice"]
            if choice == new_result:
                winners.append(member["name"])
            else:
                losers.append(member["name"])
                player["debt"] = player.get("debt", 0) + fee
        else:
            no_vote.append(member["name"])
            player["debt"] = player.get("debt", 0) + fee

    data["matches"][match_id]["result"] = new_result
    save_data(data)

    result_label = {"home": m["home_team"], "draw": "Hòa kèo", "away": m["away_team"]}[new_result]
    hcap_txt     = format_handicap(m["handicap"], m["home_team"], m["away_team"])
    msg = (
        f"SỬA KẾT QUẢ TRẬN {match_id}\n"
        f"{m['home_team']} vs {m['away_team']}\n"
        f"Kèo: {hcap_txt}\n"
        f"Kết quả cũ: {old_result or 'chưa có'} → Mới: {result_label}\n\n"
    )
    if winners:  msg += f"THẮNG ({len(winners)}): {', '.join(winners)}\n"
    if losers:   msg += f"THUA -{fee:,}đ ({len(losers)}): {', '.join(losers)}\n"
    if no_vote:  msg += f"KHÔNG BQ -{fee:,}đ ({len(no_vote)}): {', '.join(no_vote)}\n"

    await context.bot.send_message(chat_id=GROUP_ID, text=msg)
    await update.message.reply_text("Đã sửa kết quả và tính lại nợ.")


async def cmd_editdebt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Sửa nợ thủ công cho 1 thành viên.
    /suano <tên hoặc ID> <số_tiền_mới>
    Ví dụ: /suano Văn A 150000   (đặt nợ = 150k)
            /suano Văn A 0        (xóa nợ)
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Cú pháp: /suano <tên hoặc ID> <số_tiền_nợ_mới>\n"
            "Ví dụ: /suano Văn A 150000\n"
            "        /suano 123456789 0"
        )
        return

    amount_str = context.args[-1].replace(".", "").replace(",", "")
    if not amount_str.isdigit():
        await update.message.reply_text("Số tiền phải là số nguyên.")
        return

    amount = int(amount_str)
    query  = " ".join(context.args[:-1]).strip()
    data   = load_data()
    members = data.get("members", {})

    # Tìm thành viên
    found_uid = found_name = None
    if query in members:
        found_uid  = query
        found_name = members[query]["name"]
    else:
        matches_found = [(uid, m["name"]) for uid, m in members.items()
                         if query.lower() in m["name"].lower()]
        if len(matches_found) == 1:
            found_uid, found_name = matches_found[0]
        elif len(matches_found) > 1:
            names = "\n".join([f"• {n} (ID: {u})" for u, n in matches_found])
            await update.message.reply_text(f"Tìm thấy nhiều người:\n{names}\nDùng ID để chính xác hơn.")
            return
        else:
            await update.message.reply_text(f"Không tìm thấy '{query}'.")
            return

    old_debt = data["players"].get(found_uid, {}).get("debt", 0)
    data["players"].setdefault(found_uid, {"name": found_name, "debt": 0})["debt"] = amount
    save_data(data)

    await update.message.reply_text(
        f"Đã sửa nợ của {found_name}:\n"
        f"Cũ: {old_debt:,}đ → Mới: {amount:,}đ"
    )


async def cmd_editpaid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Sửa số tiền đã đóng của thành viên.
    /suadathanhtoan <tên hoặc ID> <số_tiền_đã_đóng_mới>
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Cú pháp: /suadathanhtoan <tên hoặc ID> <số_tiền_đã_đóng>\n"
            "Ví dụ: /suadathanhtoan Văn A 100000"
        )
        return

    amount_str = context.args[-1].replace(".", "").replace(",", "")
    if not amount_str.isdigit():
        await update.message.reply_text("Số tiền phải là số nguyên.")
        return

    amount  = int(amount_str)
    query   = " ".join(context.args[:-1]).strip()
    data    = load_data()
    members = data.get("members", {})

    found_uid = found_name = None
    if query in members:
        found_uid  = query
        found_name = members[query]["name"]
    else:
        matches_found = [(uid, m["name"]) for uid, m in members.items()
                         if query.lower() in m["name"].lower()]
        if len(matches_found) == 1:
            found_uid, found_name = matches_found[0]
        elif len(matches_found) > 1:
            names = "\n".join([f"• {n} (ID: {u})" for u, n in matches_found])
            await update.message.reply_text(f"Tìm thấy nhiều người:\n{names}")
            return
        else:
            await update.message.reply_text(f"Không tìm thấy '{query}'.")
            return

    old_paid = members[found_uid].get("paid", 0)
    data["members"][found_uid]["paid"] = amount
    save_data(data)

    debt   = data["players"].get(found_uid, {}).get("debt", 0)
    remain = max(0, debt - amount)
    await update.message.reply_text(
        f"Đã sửa số tiền đã đóng của {found_name}:\n"
        f"Cũ: {old_paid:,}đ → Mới: {amount:,}đ\n"
        f"Tổng nợ: {debt:,}đ | Còn lại: {remain:,}đ"
    )


async def cmd_viewmatch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xem chi tiết 1 trận: ai bình chọn gì, kết quả, nợ.
    /xemtran <match_id>
    """
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /xemtran <match_id>")
        return

    match_id = context.args[0].upper()
    data     = load_data()
    if match_id not in data["matches"]:
        await update.message.reply_text(f"Không tìm thấy trận {match_id}.")
        return

    m       = data["matches"][match_id]
    preds   = data["predictions"].get(match_id, {})
    members = data.get("members", {})
    kickoff = datetime.fromisoformat(m["kickoff"])
    hcap    = format_handicap(m["handicap"], m["home_team"], m["away_team"])
    result  = m.get("result", "Chưa có")
    score   = f"{m.get('home_score','?')}-{m.get('away_score','?')}" if m.get("result") else "Chưa đấu"

    msg = (
        f"CHI TIẾT TRẬN {match_id}\n"
        f"{m['home_team']} vs {m['away_team']}\n"
        f"Kickoff: {kickoff.strftime('%d/%m/%Y %H:%M')}\n"
        f"Kèo: {hcap}\n"
        f"Tỷ số: {score} | Kết quả kèo: {result}\n"
        f"Vòng: {m['round']} | Phí: {FEE.get(m['round'],50000):,}đ\n"
        f"Poll: {'Đã khóa' if m['locked'] else 'Đang mở'}\n\n"
        f"BÌNH CHỌN ({len(preds)} người):\n"
    )

    choice_label = {"home": "TRÊN", "draw": "HÒA", "away": "DƯỚI"}
    for uid, pred in preds.items():
        name   = pred["name"]
        choice = choice_label.get(pred["choice"], pred["choice"])
        if m.get("result"):
            outcome = "✓" if pred["choice"] == m["result"] else "✗"
        else:
            outcome = ""
        msg += f"  {outcome} {name}: {choice}\n"

    no_vote = [m2["name"] for uid, m2 in members.items() if uid not in preds]
    if no_vote:
        msg += f"\nKHÔNG BQ: {', '.join(no_vote)}"

    await update.message.reply_text(msg)



async def cmd_listsports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem danh sách sport keys có sẵn để dùng test. /danhsachgiai"""
    if not is_admin(update.effective_user.id): return
    await update.message.reply_text("Đang lấy danh sách từ API...")
    sports = fetch_available_sports()
    soccer = [s for s in sports if "soccer" in s]
    if not soccer:
        await update.message.reply_text("Không lấy được danh sách. Kiểm tra ODDS_API_KEY.")
        return
    msg = "DANH SÁCH SPORT KEYS (Soccer):\n" + "="*30 + "\n"
    for s in soccer:
        msg += f"• {s}\n"
    msg += (
        "\nCách dùng: vào Railway Variables, thêm:\n"
        "TEST_SPORT_KEY = <sport_key_muốn_dùng>\n\n"
        "Tìm key có 'friendly' để test với trận giao hữu."
    )
    await update.message.reply_text(msg)


async def cmd_testmode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem trạng thái test mode và số trận theo từng loại. /chedo"""
    if not is_admin(update.effective_user.id): return
    data = load_data()
    matches = data.get("matches", {})

    ts_count = len([m for m in matches if m.startswith("TS")])
    wc_count = len([m for m in matches if m.startswith("WC")])
    other    = len(matches) - ts_count - wc_count

    mode_status = "BẬT (Giao hữu)" if TEST_MODE else "TẮT (World Cup)"
    sport_key   = get_sport_key()

    msg = (
        f"CHẾ ĐỘ HIỆN TẠI: {mode_status}\n"
        f"Sport key: {sport_key}\n\n"
        f"TRẬN TRONG HỆ THỐNG:\n"
        f"• TS... (test): {ts_count} trận\n"
        f"• WC... (World Cup): {wc_count} trận\n"
    )
    if other:
        msg += f"• Khác: {other} trận\n"

    if ts_count > 0 and not TEST_MODE:
        msg += "\n⚠️ Còn trận TEST trong khi đang ở chế độ WC!\nGõ /xoatrancu TS để xóa"
    if wc_count > 0 and TEST_MODE:
        msg += "\n⚠️ Còn trận WC trong khi đang ở chế độ TEST!\nGõ /xoatrancu WC để xóa"

    msg += (
        "\n\nLỆNH CHUYỂN CHẾ ĐỘ:\n"
        "/xoatrancu TS — Xóa trận test\n"
        "/xoatrancu WC — Xóa trận World Cup\n"
        "/resetwc confirm — Reset hẳn sang WC\n"
    )
    await update.message.reply_text(msg)


async def cmd_clearmatches(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xóa trận theo chế độ hiện tại.
    /xoatrancu       — xóa trận theo prefix hiện tại (TS hoặc WC)
    /xoatrancu all   — xóa tất cả trận
    /xoatrancu TS    — xóa trận test (TS001, TS002...)
    /xoatrancu WC    — xóa trận World Cup (WC001, WC002...)
    """
    if not is_admin(update.effective_user.id): return

    arg = context.args[0].upper() if context.args else ("TS" if TEST_MODE else "WC")
    data = load_data()
    matches = data.get("matches", {})

    if arg == "ALL":
        to_delete = list(matches.keys())
    else:
        to_delete = [mid for mid in matches if mid.startswith(arg)]

    if not to_delete:
        await update.message.reply_text(
            f"Không có trận nào có prefix '{arg}' để xóa."
        )
        return

    # Xóa bình chọn liên quan
    for mid in to_delete:
        data["matches"].pop(mid, None)
        data["predictions"].pop(mid, None)

    save_data(data)
    await update.message.reply_text(
        f"Đã xóa {len(to_delete)} trận: {', '.join(to_delete[:10])}{'...' if len(to_delete) > 10 else ''}\n"
        f"Bình chọn liên quan cũng đã xóa.\n\n"
        f"Lưu ý: nợ tiền KHÔNG reset. Dùng /resetwc để reset cả nợ."
    )



async def cmd_resetforwc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xóa toàn bộ dữ liệu test, chuẩn bị cho World Cup thật.
    /resetwc confirm
    """
    if not is_admin(update.effective_user.id): return
    if not context.args or context.args[0] != "confirm":
        await update.message.reply_text(
            "Lệnh này sẽ:\n"
            "1. Xóa toàn bộ trận test (giao hữu)\n"
            "2. Xóa toàn bộ bình chọn test\n"
            "3. Reset nợ về 0 cho tất cả thành viên\n"
            "4. GIỮ LẠI danh sách thành viên\n\n"
            "Nếu chắc chắn, gõ:\n/resetwc confirm"
        )
        return

    data = load_data()

    # Giữ lại members, reset debt và paid về 0
    members = data.get("members", {})
    players = {}
    for uid, m in members.items():
        players[uid] = {"name": m["name"], "debt": 0}
        members[uid]["paid"] = 0

    # Xóa trận và bình chọn
    new_data = {
        "matches":     {},
        "predictions": {},
        "players":     players,
        "members":     members,
    }
    save_data(new_data)

    total = len(members)
    await update.message.reply_text(
        f"Đã reset xong! Sẵn sàng cho World Cup 2026\n\n"
        f"Giữ lại: {total} thành viên (nợ reset về 0)\n"
        f"Đã xóa: tất cả trận giao hữu + bình chọn\n\n"
        f"Bước tiếp theo:\n"
        f"1. Tắt TEST_MODE trong Railway Variables\n"
        f"2. Gõ /laykeo để lấy lịch World Cup"
    )
    await context.bot.send_message(
        chat_id=GROUP_ID,
        text="THÔNG BÁO: Hệ thống đã reset, sẵn sàng cho World Cup 2026!\n"
             "Các thành viên không cần /thamgia lại.\nChờ lịch thi đấu chính thức nhé!"
    )



async def cmd_updateodds(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Cập nhật kèo mới nhất từ nhà cái cho các trận chưa diễn ra.
    /capnhatkeo
    Dùng khi kèo thay đổi sau khi đã tạo trận.
    """
    if not is_admin(update.effective_user.id): return
    await update.message.reply_text("Đang lấy kèo mới nhất từ nhà cái...")

    games = fetch_odds("spreads")
    if not games:
        await update.message.reply_text("Không lấy được dữ liệu. Kiểm tra ODDS_API_KEY.")
        return

    data     = load_data()
    now      = datetime.now(TZ)
    updated  = 0
    skipped  = 0

    # Tạo map tên đội → kèo mới để tra nhanh
    new_odds = {}
    for g in games:
        p = parse_asian_handicap(g)
        if not p: continue
        key = (p["home_team"].lower(), p["away_team"].lower())
        new_odds[key] = p

    for match_id, m in data["matches"].items():
        # Bỏ qua trận đã có kết quả hoặc đã kickoff
        if m.get("result"): continue
        kickoff = datetime.fromisoformat(m["kickoff"])
        if kickoff <= now: continue

        key = (m["home_team"].lower(), m["away_team"].lower())
        p   = new_odds.get(key)
        if not p:
            skipped += 1
            continue

        h_new    = normalize_handicap(p["handicap"])
        changed  = []
        if abs(h_new - m.get("handicap", 0)) > 0.01:
            changed.append(f"kèo {m['handicap']} → {h_new}")
            data["matches"][match_id]["handicap"]       = h_new
            data["matches"][match_id]["has_draw_option"] = has_draw_option(h_new)
        if abs(p["home_odds"] - m.get("home_odds", 0)) > 0.01:
            changed.append(f"trên {m['home_odds']} → {p['home_odds']}")
            data["matches"][match_id]["home_odds"] = p["home_odds"]
        if abs(p["away_odds"] - m.get("away_odds", 0)) > 0.01:
            changed.append(f"dưới {m['away_odds']} → {p['away_odds']}")
            data["matches"][match_id]["away_odds"] = p["away_odds"]

        if changed:
            data["matches"][match_id]["bookmaker"] = p["bookmaker"]
            updated += 1
        else:
            skipped += 1

    save_data(data)

    msg = (
        f"Cập nhật kèo hoàn tất!\n"
        f"Đã cập nhật: {updated} trận\n"
        f"Không thay đổi: {skipped} trận\n\n"
    )
    if updated > 0:
        msg += "Lưu ý: Nếu poll đã gửi, kèo hiển thị trong poll KHÔNG tự thay đổi.\n"
        msg += "Dùng /khoabinhchon rồi /guibinhchon để gửi lại poll với kèo mới."
    await update.message.reply_text(msg)



async def cmd_deletematch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xóa 1 trận khỏi hệ thống.
    /xoatran <id>
    """
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /xoatran <id>\nVD: /xoatran TS001")
        return

    match_id = context.args[0].upper()
    data = load_data()
    if match_id not in data["matches"]:
        await update.message.reply_text(f"Không tìm thấy trận {match_id}.")
        return

    m = data["matches"][match_id]
    # Bỏ ghim poll nếu đang ghim
    if m.get("poll_message_id") and not m.get("locked"):
        try:
            await context.bot.unpin_chat_message(
                chat_id=GROUP_ID,
                message_id=m["poll_message_id"]
            )
        except Exception:
            pass

    # Xóa trận + bình chọn liên quan
    del data["matches"][match_id]
    data["predictions"].pop(match_id, None)
    save_data(data)

    await update.message.reply_text(
        f"Đã xóa trận {match_id}: {m['home_team']} vs {m['away_team']}\n"
        f"Bình chọn liên quan cũng đã xóa.\n"
        f"Lưu ý: nợ tiền KHÔNG bị hoàn lại."
    )


async def cmd_editmatch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Chỉnh sửa thông tin trận đấu.
    /suatran <id> <field> <giá trị mới>
    Field: kickoff | handicap | home_odds | away_odds | round | home_team | away_team
    VD: /suatran TS001 kickoff 2026-06-10 02:00
        /suatran TS001 handicap -0.75
        /suatran TS001 home_odds 1.85
        /suatran TS001 round knockout
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 3:
        await update.message.reply_text(
            "Cú pháp: /suatran <id> <field> <giá trị>\n\n"
            "Các field có thể sửa:\n"
            "• kickoff — VD: 2026-06-10 02:00\n"
            "• handicap — VD: -0.75\n"
            "• home_odds — VD: 1.85\n"
            "• away_odds — VD: 1.95\n"
            "• round — VD: group / knockout / final\n"
            "• home_team — VD: Brazil\n"
            "• away_team — VD: Argentina"
        )
        return

    match_id = context.args[0].upper()
    field    = context.args[1].lower()
    value    = " ".join(context.args[2:])
    data     = load_data()

    if match_id not in data["matches"]:
        await update.message.reply_text(f"Không tìm thấy trận {match_id}.")
        return

    m       = data["matches"][match_id]
    allowed = ["kickoff", "handicap", "home_odds", "away_odds", "round", "home_team", "away_team"]

    if field not in allowed:
        await update.message.reply_text(
            f"Field không hợp lệ. Chọn một trong:\n{', '.join(allowed)}"
        )
        return

    old_val = m.get(field, "?")
    try:
        if field == "kickoff":
            # Parse ngày giờ
            kickoff = TZ.localize(datetime.strptime(value, "%Y-%m-%d %H:%M"))
            data["matches"][match_id]["kickoff"] = kickoff.isoformat()
            # Cập nhật lịch tự động
            scheduler = context.bot_data.get("scheduler")
            if scheduler:
                now = datetime.now(TZ)
                send_time = kickoff - timedelta(hours=POLL_BEFORE)
                if send_time > now and not m.get("poll_message_id"):
                    scheduler.add_job(auto_send_poll_job, "date", run_date=send_time,
                        args=[match_id, context.application],
                        id=f"send_{match_id}", replace_existing=True)
                if kickoff > now and not m.get("locked"):
                    scheduler.add_job(lock_poll_job, "date", run_date=kickoff,
                        args=[match_id, context.application],
                        id=f"lock_{match_id}", replace_existing=True)
            new_val = kickoff.strftime("%d/%m/%Y %H:%M")
        elif field in ["handicap", "home_odds", "away_odds"]:
            new_val = float(value)
            data["matches"][match_id][field] = new_val
            if field == "handicap":
                data["matches"][match_id]["handicap"] = normalize_handicap(new_val)
                data["matches"][match_id]["has_draw_option"] = has_draw_option(new_val)
                new_val = normalize_handicap(new_val)
        elif field == "round":
            if value not in ["group", "knockout", "final"]:
                await update.message.reply_text("round phải là: group / knockout / final")
                return
            data["matches"][match_id]["round"] = value
            new_val = value
        else:
            data["matches"][match_id][field] = value
            new_val = value

        save_data(data)
        await update.message.reply_text(
            f"Đã cập nhật trận {match_id}:\n"
            f"{field}: {old_val} → {new_val}\n\n"
            f"Dùng /xemtran {match_id} để xem lại chi tiết."
        )
    except ValueError as e:
        await update.message.reply_text(f"Giá trị không hợp lệ: {e}")



async def cmd_resetdata(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xóa toàn bộ dữ liệu test, bắt đầu lại từ đầu.
    /xoahetdata confirm
    """
    if not is_admin(update.effective_user.id): return
    if not context.args or context.args[0] != "confirm":
        await update.message.reply_text(
            "CẢNH BÁO: Lệnh này xóa TOÀN BỘ dữ liệu!\n"
            "Bao gồm: thành viên, trận đấu, bình chọn, nợ tiền.\n\n"
            "Nếu chắc chắn, gõ:\n/xoahetdata confirm"
        )
        return
    # Xóa sạch
    empty = {"matches": {}, "predictions": {}, "players": {}, "members": {}}
    save_data(empty)
    await update.message.reply_text(
        "Đã xóa toàn bộ dữ liệu!\n"
        "Bot sẵn sàng cho World Cup 2026.\n\n"
        "Bắt đầu bằng:\n"
        "1. /laykeo — lấy lịch thi đấu\n"
        "2. Nhờ thành viên gõ /thamgia trong nhóm"
    )


async def cmd_syncdata(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sync data lên GitHub thủ công. /luudata"""
    if not is_admin(update.effective_user.id): return
    if not GITHUB_TOKEN:
        await update.message.reply_text("Chưa có GITHUB_TOKEN trong Variables.")
        return
    push_data_to_github()
    await update.message.reply_text("Đã sync data.json lên GitHub!")


# ============================================================
#  KHỞI ĐỘNG
# ============================================================

async def post_init(application):
    # Khôi phục data từ GitHub trước khi làm gì
    pull_data_from_github()

    scheduler = AsyncIOScheduler(timezone=TZ)
    scheduler.start()
    application.bot_data["scheduler"] = scheduler

    data = load_data()
    now  = datetime.now(TZ)
    restored = 0

    for match_id, m in data["matches"].items():
        kickoff = datetime.fromisoformat(m["kickoff"])

        # Khôi phục lịch gửi poll
        if not m.get("poll_message_id") and kickoff > now:
            send_time = kickoff - timedelta(hours=POLL_BEFORE)
            if send_time > now:
                scheduler.add_job(auto_send_poll_job, "date", run_date=send_time,
                    args=[match_id, application], id=f"send_{match_id}", replace_existing=True)
                restored += 1

        # Khôi phục lịch khóa poll
        if not m["locked"] and m.get("poll_message_id") and kickoff > now:
            scheduler.add_job(lock_poll_job, "date", run_date=kickoff,
                args=[match_id, application], id=f"lock_{match_id}", replace_existing=True)
            restored += 1

        # Khôi phục lịch nhắc nhở trước 1 tiếng
        if not m.get("locked") and not m.get("result") and kickoff > now:
            remind_time = kickoff - timedelta(hours=1)
            if remind_time > now:
                scheduler.add_job(remind_unvoted_job, "date", run_date=remind_time,
                    args=[match_id, application], id=f"remind_{match_id}", replace_existing=True)

        # Khôi phục lịch lấy kết quả tự động
        if not m.get("result") and kickoff < now:
            result_time = kickoff + timedelta(hours=2, minutes=30)
            if result_time > now:
                scheduler.add_job(auto_fetch_result_job, "date", run_date=result_time,
                    args=[match_id, application], id=f"result_{match_id}", replace_existing=True)
                restored += 1
            elif (now - kickoff).total_seconds() < 6 * 3600:
                # Kickoff đã qua nhưng chưa quá 6 tiếng → thử lấy kết quả ngay
                scheduler.add_job(auto_fetch_result_job, "date",
                    run_date=now + timedelta(minutes=1),
                    args=[match_id, application], id=f"result_now_{match_id}", replace_existing=True)
                restored += 1

    logger.info(f"Bot đã khởi động! Khôi phục {restored} lịch.")


# ============================================================
#  LẤY KẾT QUẢ TỪ API-FOOTBALL
# ============================================================

# Bảng ánh xạ tên đội Anh (Odds API) → tên trên API-Football nếu khác nhau
TEAM_NAME_MAP = {
    "south korea": "korea republic",
    "north korea": "korea dprk",
    "usa": "united states",
    "ivory coast": "côte d'ivoire",
    "cape verde": "cape verde islands",
    "czech republic": "czechia",
}

def _normalize_team(name: str) -> str:
    n = name.lower().strip()
    return TEAM_NAME_MAP.get(n, n)

def _team_match(name1: str, name2: str) -> bool:
    """So khớp tên 2 đội linh hoạt."""
    a = _normalize_team(name1)
    b = _normalize_team(name2)
    return a in b or b in a or a == b

def fetch_match_result_api(home_team: str, away_team: str, match_date: str) -> dict | None:
    """
    Lấy kết quả trận đấu từ API-Football (league=1, season=2026).
    Trả về {"home_score", "away_score"} hoặc None.
    Lấy theo NGÀY rồi so khớp tên đội (đội nhà/khách có thể đảo).
    """
    if not APIFOOTBALL_KEY:
        logger.warning("Chưa cấu hình APIFOOTBALL_KEY")
        return None
    try:
        url = "https://v3.football.api-sports.io/fixtures"
        headers = {"x-apisports-key": APIFOOTBALL_KEY}
        # Lấy TẤT CẢ trận trong ngày (không lọc status để bắt cả FT/AET/PEN)
        params  = {"league": "1", "season": "2026", "date": match_date}
        r = requests.get(url, headers=headers, params=params, timeout=15)

        if r.status_code != 200:
            logger.error(f"API-Football HTTP {r.status_code}: {r.text[:150]}")
            return None

        body = r.json()
        # API-Football trả lỗi trong field "errors"
        if body.get("errors"):
            logger.error(f"API-Football errors: {body['errors']}")
            return None

        fixtures = body.get("response", [])
        logger.info(f"API-Football: {len(fixtures)} trận ngày {match_date}")

        for fix in fixtures:
            h = fix["teams"]["home"]["name"]
            a = fix["teams"]["away"]["name"]
            status = fix["fixture"]["status"]["short"]  # FT, AET, PEN, NS, ...

            # So khớp 2 chiều (phòng trường hợp đảo sân)
            match_direct  = _team_match(home_team, h) and _team_match(away_team, a)
            match_swapped = _team_match(home_team, a) and _team_match(away_team, h)

            if match_direct or match_swapped:
                # Chỉ lấy khi trận đã kết thúc
                if status not in ["FT", "AET", "PEN"]:
                    logger.info(f"Trận {h} vs {a} chưa xong (status={status})")
                    return None
                goals = fix["goals"]
                if goals["home"] is None or goals["away"] is None:
                    return None
                # Nếu bị đảo sân, đảo lại tỷ số cho đúng đội nhà của ta
                if match_swapped and not match_direct:
                    return {"home_score": goals["away"], "away_score": goals["home"]}
                return {"home_score": goals["home"], "away_score": goals["away"]}

        logger.warning(f"Không tìm thấy trận {home_team} vs {away_team} ngày {match_date}")
        return None
    except Exception as e:
        logger.error(f"fetch_match_result_api lỗi: {e}")
        return None


def determine_keo_result(home_score: int, away_score: int, handicap: float) -> str:
    """
    Tính kết quả kèo châu Á từ tỷ số thực tế.

    Quy ước handicap (theo nhà cái):
      handicap < 0: đội nhà chấp |handicap| trái cho đội khách
        VD: -0.5 → nhà chấp 0.5, nhà phải thắng mới thắng kèo
        VD: -1.0 → nhà chấp 1, nhà phải thắng 2+ mới thắng kèo
      handicap > 0: đội khách chấp handicap trái cho đội nhà
        VD: +0.5 → khách chấp 0.5, khách phải thắng mới thắng kèo
        VD: +1.0 → khách chấp 1, khách phải thắng 2+ mới thắng kèo
      handicap == 0: kèo chẵn

    Cách tính: lấy hiệu tỷ số (home - away) so với mức chấp
      diff = home_score - away_score
      Nếu handicap < 0 (nhà chấp):  diff > |handicap| → nhà thắng kèo
      Nếu handicap > 0 (khách chấp): diff < -handicap → khách thắng kèo
    """
    h    = normalize_handicap(handicap)
    diff = home_score - away_score  # dương: nhà dẫn, âm: khách dẫn

    # Áp kèo: điều chỉnh diff theo handicap
    # h âm (nhà chấp): nhà bị trừ → adjusted = diff + h (h âm → làm khó nhà)
    # h dương (khách chấp): nhà được cộng → adjusted = diff + h
    adjusted = diff + h

    if adjusted > 0:   return "home"
    elif adjusted < 0: return "away"
    else:              return "draw"  # chỉ xảy ra với kèo nguyên



async def auto_fetch_result_job(match_id: str, app):
    """
    Tự động lấy kết quả 1 trận cụ thể sau khi kết thúc.
    Nếu chưa có kết quả thì thử lại sau 30 phút (tối đa 3 lần).
    """
    data = load_data()
    if match_id not in data["matches"]:
        return
    m = data["matches"][match_id]
    if m.get("result"):
        return  # Đã có kết quả rồi

    if not APIFOOTBALL_KEY:
        logger.warning(f"Không có APIFOOTBALL_KEY, bỏ qua auto fetch trận {match_id}")
        return

    # Test mode dùng giao hữu - API-Football league khác, không tự lấy được
    if TEST_MODE:
        logger.info(f"TEST MODE: bỏ qua auto fetch trận {match_id}, nhập thủ công bằng /ketqua")
        try:
            await app.bot.send_message(
                chat_id=ADMIN_ID,
                text=f"[TEST] Trận {match_id} đã kết thúc.\n"
                     f"{m['home_team']} vs {m['away_team']}\n"
                     f"Nhập kết quả thủ công: /ketqua {match_id} <home|draw|away>"
            )
        except Exception:
            pass
        return

    kickoff    = datetime.fromisoformat(m["kickoff"])
    match_date = kickoff.strftime("%Y-%m-%d")
    score      = fetch_match_result_api(m["home_team"], m["away_team"], match_date)

    if not score:
        # Thử lại sau 30 phút nếu chưa có kết quả
        retry_count = m.get("result_retry", 0)
        if retry_count < 3:
            data["matches"][match_id]["result_retry"] = retry_count + 1
            save_data(data)
            scheduler = app.bot_data.get("scheduler")
            if scheduler:
                retry_time = datetime.now(TZ) + timedelta(minutes=30)
                scheduler.add_job(auto_fetch_result_job, "date",
                    run_date=retry_time,
                    args=[match_id, app],
                    id=f"result_retry_{match_id}_{retry_count}",
                    replace_existing=True)
            logger.info(f"Chưa có kết quả trận {match_id}, thử lại sau 30 phút (lần {retry_count+1}/3)")
        else:
            logger.warning(f"Không tìm được kết quả trận {match_id} sau 3 lần thử")
            await app.bot.send_message(
                chat_id=ADMIN_ID,
                text=f"Không tự động lấy được kết quả trận {match_id}\n"
                     f"{m['home_team']} vs {m['away_team']}\n"
                     f"Hãy nhập thủ công bằng /ketqua {match_id} <home|draw|away>"
            )
        return

    # Có kết quả — tính thắng thua
    keo_result = determine_keo_result(score["home_score"], score["away_score"], m["handicap"])
    data = load_data()  # load lại đề phòng có thay đổi
    data["matches"][match_id]["result"]     = keo_result
    data["matches"][match_id]["home_score"] = score["home_score"]
    data["matches"][match_id]["away_score"] = score["away_score"]

    fee     = FEE.get(m["round"], 50000)
    members = data.get("members", {})
    preds   = data["predictions"].get(match_id, {})
    winners, losers, no_vote = [], [], []

    for uid, member in members.items():
        player = data["players"].setdefault(uid, {"name": member["name"], "debt": 0})
        player["name"] = member["name"]
        if uid in preds:
            choice = preds[uid]["choice"]
            if choice == keo_result:
                winners.append(member["name"])
            else:
                losers.append(member["name"])
                player["debt"] = player.get("debt", 0) + fee
        else:
            # Không bình chọn = thua
            no_vote.append(member["name"])
            player["debt"] = player.get("debt", 0) + fee

    save_data(data)

    # Thông báo vào nhóm
    hcap_txt     = format_handicap(m["handicap"], m["home_team"], m["away_team"])
    result_label = {"home": m["home_team"], "draw": "Hòa kèo", "away": m["away_team"]}[keo_result]
    msg = (
        f"KẾT QUẢ TỰ ĐỘNG - Trận {match_id}\n"
        f"{m['home_team']} {score['home_score']} - {score['away_score']} {m['away_team']}\n"
        f"Kèo: {hcap_txt} → {result_label} thắng kèo\n\n"
    )
    if winners:  msg += f"THẮNG ({len(winners)}): {', '.join(winners)}\n"
    if losers:   msg += f"THUA -{fee:,}đ ({len(losers)}): {', '.join(losers)}\n"
    if no_vote:  msg += f"KHÔNG BQ -{fee:,}đ ({len(no_vote)}): {', '.join(no_vote)}\n"

    try:
        await app.bot.send_message(chat_id=GROUP_ID, text=msg)
        logger.info(f"Đã tự động cập nhật kết quả trận {match_id}: {keo_result}")
    except Exception as e:
        logger.error(f"Lỗi gửi kết quả tự động {match_id}: {e}")


async def cmd_testapi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Kiểm tra API-Football trả về gì cho 1 ngày.
    /testapi <YYYY-MM-DD>
    VD: /testapi 2026-06-12
    """
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /testapi <YYYY-MM-DD>\nVD: /testapi 2026-06-12")
        return

    match_date = context.args[0]
    if not APIFOOTBALL_KEY:
        await update.message.reply_text("Chưa có APIFOOTBALL_KEY trong Variables!")
        return

    await update.message.reply_text(f"Đang kiểm tra API-Football ngày {match_date}...")
    try:
        url = "https://v3.football.api-sports.io/fixtures"
        headers = {"x-apisports-key": APIFOOTBALL_KEY}
        params  = {"league": "1", "season": "2026", "date": match_date}
        r = requests.get(url, headers=headers, params=params, timeout=15)

        body = r.json()
        # Kiểm tra quota
        remaining = r.headers.get("x-ratelimit-requests-remaining", "?")

        if body.get("errors"):
            await update.message.reply_text(
                f"API trả về LỖI:\n{body['errors']}\n\n"
                f"Thường gặp: hết quota, sai key, hoặc plan free không hỗ trợ."
            )
            return

        fixtures = body.get("response", [])
        if not fixtures:
            await update.message.reply_text(
                f"API hoạt động nhưng KHÔNG có trận nào ngày {match_date}.\n"
                f"Quota còn: {remaining}\n\n"
                f"Có thể: ngày sai, hoặc plan free chưa có dữ liệu WC 2026."
            )
            return

        msg = f"API OK! Có {len(fixtures)} trận ngày {match_date}:\n"
        msg += f"Quota còn: {remaining}\n" + "="*28 + "\n"
        for fix in fixtures[:8]:
            h = fix["teams"]["home"]["name"]
            a = fix["teams"]["away"]["name"]
            status = fix["fixture"]["status"]["short"]
            gh = fix["goals"]["home"]
            ga = fix["goals"]["away"]
            score = f"{gh}-{ga}" if gh is not None else "chưa đá"
            msg += f"• {h} vs {a}: {score} ({status})\n"

        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"Lỗi: {e}")



async def cmd_fetchresult(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Tự động lấy kết quả từ API-Football cho tất cả trận đã kết thúc.
    /laykequa
    """
    if not is_admin(update.effective_user.id): return
    if not APIFOOTBALL_KEY:
        await update.message.reply_text("Chưa có APIFOOTBALL_KEY trong file .env!")
        return

    await update.message.reply_text("Đang lấy kết quả từ API-Football...")
    data = load_data()
    now  = datetime.now(TZ)
    updated = 0
    errors  = 0

    for match_id, m in data["matches"].items():
        if m.get("result"):
            continue  # Đã có kết quả rồi
        kickoff = datetime.fromisoformat(m["kickoff"])
        # Chỉ xử lý trận đã kết thúc hơn 2 tiếng
        if now < kickoff + timedelta(hours=2):
            continue

        match_date = kickoff.strftime("%Y-%m-%d")
        score = fetch_match_result_api(m["home_team"], m["away_team"], match_date)
        if not score:
            errors += 1
            continue

        keo_result = determine_keo_result(score["home_score"], score["away_score"], m["handicap"])
        data["matches"][match_id]["result"]     = keo_result
        data["matches"][match_id]["home_score"] = score["home_score"]
        data["matches"][match_id]["away_score"] = score["away_score"]

        # Tính thắng/thua cho tất cả thành viên
        fee       = FEE.get(m["round"], 50000)
        members   = data.get("members", {})
        preds     = data["predictions"].get(match_id, {})
        winners, losers, no_vote = [], [], []

        for uid, member in members.items():
            if uid in preds:
                choice = preds[uid]["choice"]
                player = data["players"].setdefault(uid, {"name": member["name"], "debt": 0})
                player["name"] = member["name"]
                if choice == keo_result:
                    winners.append(member["name"])
                else:
                    losers.append(member["name"])
                    player["debt"] = player.get("debt", 0) + fee
            else:
                # Không bình chọn = thua cuộc
                player = data["players"].setdefault(uid, {"name": member["name"], "debt": 0})
                player["name"] = member["name"]
                no_vote.append(member["name"])
                player["debt"] = player.get("debt", 0) + fee

        save_data(data)
        updated += 1

        # Thông báo kết quả vào nhóm
        hcap_txt     = format_handicap(m["handicap"], m["home_team"], m["away_team"])
        result_label = {"home": m["home_team"], "draw": "Hòa kèo", "away": m["away_team"]}[keo_result]
        msg = (
            f"KẾT QUẢ TRẬN {match_id}\n"
            f"{m['home_team']} {score['home_score']} - {score['away_score']} {m['away_team']}\n"
            f"Kèo: {hcap_txt} → {result_label} thắng kèo\n\n"
        )
        if winners:  msg += f"THẮNG ({len(winners)}): {', '.join(winners)}\n"
        if losers:   msg += f"THUA -{fee:,}đ ({len(losers)}): {', '.join(losers)}\n"
        if no_vote:  msg += f"KHÔNG BQ -{fee:,}đ ({len(no_vote)}): {', '.join(no_vote)}\n"

        try:
            await context.bot.send_message(chat_id=GROUP_ID, text=msg)
        except Exception as e:
            logger.error(f"Lỗi gửi kết quả {match_id}: {e}")

    await update.message.reply_text(
        f"Đã cập nhật {updated} trận.\n"
        f"Không tìm thấy kết quả: {errors} trận.\n"
        f"Gõ /xuatfile để xuất file tính tiền."
    )


# ============================================================
#  QUẢN LÝ THÀNH VIÊN
# ============================================================


async def cmd_rules(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Gửi thể lệ vào nhóm. /guithele"""
    if not is_admin(update.effective_user.id): return
    await context.bot.send_message(
        chat_id=GROUP_ID,
        text=RULES_TEXT,
        parse_mode=RULES_PARSE_MODE
    )
    await update.message.reply_text("Đã gửi thể lệ vào nhóm!")


async def cmd_myrules(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem thể lệ cá nhân. /thele"""
    await update.message.reply_text(
        RULES_TEXT,
        parse_mode=RULES_PARSE_MODE
    )



async def cmd_join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Thành viên tự đăng ký tham gia. Gõ /thamgia trong nhóm."""
    user    = update.effective_user
    user_id = str(user.id)
    name    = user.full_name
    data    = load_data()

    if user_id in data.get("members", {}):
        await update.message.reply_text(
            f"Bạn đã đăng ký rồi: {name}\nGõ /lichsu để xem lịch sử bình chọn."
        )
        return

    data.setdefault("members", {})[user_id] = {"name": name, "paid": 0}
    data["players"].setdefault(user_id, {"name": name, "debt": 0})["name"] = name
    save_data(data)

    # Đếm số thành viên
    total = len(data["members"])
    await update.message.reply_text(
        f"Đã đăng ký tham gia!\n"
        f"Tên: {name}\n"
        f"Tổng thành viên: {total} người\n\n"
        f"Bạn sẽ bị tính thua nếu không bình chọn trước giờ kickoff."
    )
    # Gửi thể lệ riêng cho người mới
    try:
        await context.bot.send_message(
            chat_id=user.id,
            text=RULES_TEXT,
            parse_mode=RULES_PARSE_MODE
        )
    except Exception:
        # Nếu không nhắn riêng được (chưa từng chat với bot)
        pass
    # Thông báo cho admin
    try:
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"Thành viên mới đăng ký:\n{name} (ID: {user_id})\nTổng: {total} người"
        )
    except Exception:
        pass


async def cmd_members(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem danh sách thành viên đã đăng ký. /danhsach"""
    if not is_admin(update.effective_user.id): return
    data    = load_data()
    members = data.get("members", {})
    if not members:
        await update.message.reply_text(
            "Chưa có thành viên nào.\n"
            "Nhờ mọi người gõ /thamgia trong nhóm để đăng ký."
        )
        return
    msg = f"DANH SÁCH THÀNH VIÊN ({len(members)} người)\n" + "="*28 + "\n"
    for i, (uid, m) in enumerate(members.items(), 1):
        debt   = data["players"].get(uid, {}).get("debt", 0)
        paid   = m.get("paid", 0)
        remain = max(0, debt - paid)
        status = f"Còn nợ {remain:,}đ" if remain > 0 else "Đã đóng đủ"
        msg   += f"{i}. {m['name']}\n    ID: {uid} | {status}\n"
    msg += "\nDùng tên hoặc ID cho các lệnh /dathanhtoan, /suano..."
    await update.message.reply_text(msg)


async def cmd_removemember(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xóa thành viên khỏi danh sách. /xoathanhvien <user_id>"""
    if not is_admin(update.effective_user.id): return
    if not context.args:
        await update.message.reply_text("Cú pháp: /xoathanhvien <user_id>")
        return
    uid  = str(context.args[0])
    data = load_data()
    if uid not in data.get("members", {}):
        await update.message.reply_text("Không tìm thấy thành viên này.")
        return
    name = data["members"][uid]["name"]
    del data["members"][uid]
    save_data(data)
    await update.message.reply_text(f"Đã xóa {name} khỏi danh sách.")


async def cmd_addmember(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Thêm thành viên vào danh sách tham gia.
    /themthanhvien @username TênHiển_Thị
    hoặc /themthanhvien 123456789 Tên
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Cú pháp: /themthanhvien <user_id> <Tên>\n"
            "Ví dụ: /themthanhvien 123456789 Nguyễn Văn A\n\n"
            "Lấy user_id: nhờ thành viên nhắn tin cho @userinfobot"
        )
        return
    try:
        uid  = str(context.args[0])
        name = " ".join(context.args[1:])
        data = load_data()
        data.setdefault("members", {})[uid] = {"name": name, "paid": 0}
        data["players"].setdefault(uid, {"name": name, "debt": 0})["name"] = name
        save_data(data)
        await update.message.reply_text(f"Đã thêm thành viên: {name} (ID: {uid})")
    except Exception as e:
        await update.message.reply_text(f"Lỗi: {e}")


async def cmd_paid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Xác nhận thành viên đã đóng tiền.
    /dathanhtoan <tên hoặc user_id> <số_tiền>
    Ví dụ: /dathanhtoan Nguyễn Văn A 150000
            /dathanhtoan 123456789 150000
    """
    if not is_admin(update.effective_user.id): return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Cú pháp: /dathanhtoan <tên hoặc ID> <số_tiền>\n"
            "Ví dụ: /dathanhtoan Nguyễn Văn A 150000\n"
            "        /dathanhtoan 123456789 150000"
        )
        return
    try:
        # Số cuối luôn là số tiền
        amount_str = context.args[-1]
        # Loại bỏ dấu chấm/phẩy nếu có (ví dụ: 150.000 hoặc 150,000)
        amount_str = amount_str.replace(".", "").replace(",", "")
        if not amount_str.isdigit():
            await update.message.reply_text(
                "Số tiền phải là số nguyên ở cuối lệnh.\n"
                "Ví dụ: /dathanhtoan Văn A 150000\n"
                "        /dathanhtoan Trần Thị B 100000"
            )
            return
        amount = int(amount_str)
        query  = " ".join(context.args[:-1]).strip()
        data      = load_data()
        members   = data.get("members", {})

        # Tìm thành viên theo ID hoặc tên
        found_uid  = None
        found_name = None

        # Thử tìm theo ID trước
        if query in members:
            found_uid  = query
            found_name = members[query]["name"]
        else:
            # Tìm theo tên (không phân biệt hoa thường)
            query_lower = query.lower()
            matches_found = []
            for uid, m in members.items():
                if query_lower in m["name"].lower():
                    matches_found.append((uid, m["name"]))

            if len(matches_found) == 1:
                found_uid, found_name = matches_found[0]
            elif len(matches_found) > 1:
                names = "\n".join([f"• {n} (ID: {u})" for u, n in matches_found])
                await update.message.reply_text(
                    f"Tìm thấy {len(matches_found)} người tên gần giống:\n{names}\n\n"
                    "Dùng ID để xác nhận chính xác hơn:\n"
                    f"/dathanhtoan <ID> {amount}"
                )
                return
            else:
                # Gợi ý danh sách thành viên
                all_names = "\n".join([f"• {m['name']} (ID: {u})" for u, m in members.items()])
                await update.message.reply_text(
                    f"Không tìm thấy '{query}'.\n\n"
                    f"Danh sách thành viên:\n{all_names}"
                )
                return

        # Cập nhật tiền đã đóng
        data["members"][found_uid]["paid"] = data["members"][found_uid].get("paid", 0) + amount
        player = data["players"].setdefault(found_uid, {"name": found_name, "debt": 0})
        debt   = player.get("debt", 0)
        paid   = data["members"][found_uid]["paid"]
        remain = max(0, debt - paid)
        save_data(data)

        status = "Đã đóng đủ!" if remain == 0 else f"Còn thiếu: {remain:,}đ"
        await update.message.reply_text(
            f"Đã ghi nhận {found_name} đóng {amount:,}đ\n"
            f"Tổng nợ: {debt:,}đ | Đã đóng: {paid:,}đ\n"
            f"{status}"
        )
    except Exception as e:
        await update.message.reply_text(f"Lỗi: {e}")


async def cmd_unpaid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xem danh sách chưa đóng đủ tiền. /conno"""
    if not is_admin(update.effective_user.id): return
    data    = load_data()
    members = data.get("members", {})
    if not members:
        await update.message.reply_text("Chưa có thành viên nào. Dùng /themthanhvien để thêm.")
        return

    msg = "DANH SÁCH CHƯA ĐÓNG ĐỦ TIỀN\n" + "="*30 + "\n"
    total_owed = 0
    for uid, member in members.items():
        player = data["players"].get(uid, {})
        debt   = player.get("debt", 0)
        paid   = member.get("paid", 0)
        remain = debt - paid
        if remain > 0:
            msg += f"• {member['name']}: nợ {debt:,}đ | đã đóng {paid:,}đ | còn {remain:,}đ\n"
            total_owed += remain
    msg += f"\nTổng chưa thu: {total_owed:,}đ"
    await update.message.reply_text(msg)


# ============================================================
#  XUẤT EXCEL
# ============================================================

async def cmd_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xuất file Excel tính tiền cho tất cả thành viên. /xuatfile"""
    if not is_admin(update.effective_user.id): return

    data    = load_data()
    members = data.get("members", {})
    matches = data.get("matches", {})

    if not members:
        await update.message.reply_text("Chưa có thành viên nào. Dùng /themthanhvien để thêm.")
        return

    await update.message.reply_text("Đang tạo file Excel...")

    # Lấy danh sách trận đã có kết quả, sắp theo thời gian
    done_matches = {
        mid: m for mid, m in matches.items() if m.get("result")
    }
    done_matches = dict(sorted(done_matches.items(),
                               key=lambda x: x[1]["kickoff"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tính Tiền World Cup 2026"

    # ---- Style ----
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    GRAY   = PatternFill("solid", fgColor="D9D9D9")
    BLUE   = PatternFill("solid", fgColor="BDD7EE")
    HEADER = PatternFill("solid", fgColor="1F4E79")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # ---- Tiêu đề chính ----
    ws.merge_cells("A1:B1")
    ws["A1"] = "BẢNG TÍNH TIỀN - WORLD CUP 2026"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = HEADER
    ws["A1"].alignment = center

    total_cols = 2 + len(done_matches) + 3  # STT+Tên + các trận + Tổng nợ + Đã đóng + Còn lại
    if total_cols > 2:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].fill = HEADER

    ws["A2"] = f"Xuất lúc: {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    # ---- Header hàng 3: STT | Tên | các trận | Tổng nợ | Đã đóng | Còn lại ----
    headers = ["STT", "Họ tên"]
    for mid, m in done_matches.items():
        ko  = datetime.fromisoformat(m["kickoff"]).strftime("%d/%m")
        fee = FEE.get(m["round"], 50000)
        hdr = f"{mid}\n{m['home_team']}\nvs\n{m['away_team']}\n{ko}\n(-{fee//1000}k)"
        headers.append(hdr)
    headers += ["Tổng nợ", "Đã đóng", "Còn lại"]

    for col, hdr in enumerate(headers, 1):
        cell            = ws.cell(row=3, column=col, value=hdr)
        cell.font       = bold_white
        cell.fill       = HEADER
        cell.alignment  = center
        cell.border     = thin
        ws.column_dimensions[get_column_letter(col)].width = 14 if col > 2 else (5 if col == 1 else 22)

    ws.row_dimensions[3].height = 80

    # ---- Dữ liệu từng thành viên ----
    for row_idx, (uid, member) in enumerate(members.items(), 4):
        player    = data["players"].get(uid, {})
        preds_uid = {mid: data["predictions"].get(mid, {}).get(uid) for mid in done_matches}

        ws.cell(row=row_idx, column=1, value=row_idx - 3).alignment = center
        name_cell            = ws.cell(row=row_idx, column=2, value=member["name"])
        name_cell.font       = bold_black
        name_cell.alignment  = Alignment(vertical="center")

        # Từng trận
        for col_idx, (mid, m) in enumerate(done_matches.items(), 3):
            fee    = FEE.get(m["round"], 50000)
            result = m.get("result")
            pred   = preds_uid.get(mid)
            cell   = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border    = thin

            if pred is None:
                # Không bình chọn = thua
                cell.value = f"KBQ\n-{fee//1000}k"
                cell.fill  = YELLOW
                cell.font  = Font(color="7F6000")
            elif pred["choice"] == result:
                cell.value = "THẮNG"
                cell.fill  = GREEN
                cell.font  = Font(color="375623")
            else:
                choice_label = {"home": "TRÊN", "draw": "HÒA", "away": "DƯỚI"}.get(pred["choice"], pred["choice"])
                cell.value = f"THUA\n({choice_label})\n-{fee//1000}k"
                cell.fill  = RED
                cell.font  = Font(color="9C0006")

        # Tổng nợ / Đã đóng / Còn lại
        debt   = player.get("debt", 0)
        paid   = member.get("paid", 0)
        remain = max(0, debt - paid)
        total_col  = 3 + len(done_matches)
        paid_col   = total_col + 1
        remain_col = total_col + 2

        debt_cell          = ws.cell(row=row_idx, column=total_col,  value=debt)
        debt_cell.number_format = '#,##0"đ"'
        debt_cell.font     = bold_black
        debt_cell.fill     = RED if debt > 0 else GREEN
        debt_cell.alignment = center
        debt_cell.border   = thin

        paid_cell          = ws.cell(row=row_idx, column=paid_col,   value=paid)
        paid_cell.number_format = '#,##0"đ"'
        paid_cell.fill     = BLUE
        paid_cell.alignment = center
        paid_cell.border   = thin

        remain_cell        = ws.cell(row=row_idx, column=remain_col, value=remain)
        remain_cell.number_format = '#,##0"đ"'
        remain_cell.font   = bold_black
        remain_cell.fill   = RED if remain > 0 else GREEN
        remain_cell.alignment = center
        remain_cell.border = thin

        ws.row_dimensions[row_idx].height = 40

    # ---- Hàng tổng cộng ----
    last_row = 3 + len(members) + 1
    ws.cell(row=last_row, column=1, value="TỔNG").font = bold_black
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)

    total_col  = 3 + len(done_matches)
    paid_col   = total_col + 1
    remain_col = total_col + 2

    total_debt   = sum(data["players"].get(uid, {}).get("debt",  0) for uid in members)
    total_paid   = sum(members[uid].get("paid", 0)                  for uid in members)
    total_remain = sum(max(0, data["players"].get(uid, {}).get("debt", 0) - members[uid].get("paid", 0)) for uid in members)

    for col, val in [(total_col, total_debt), (paid_col, total_paid), (remain_col, total_remain)]:
        c = ws.cell(row=last_row, column=col, value=val)
        c.number_format = '#,##0"đ"'
        c.font  = bold_black
        c.fill  = GRAY
        c.alignment = center
        c.border = thin

    ws.cell(row=last_row, column=1).fill = GRAY
    ws.cell(row=last_row, column=2).fill = GRAY

    # ---- Chú thích màu ----
    note_row = last_row + 2
    ws.cell(row=note_row, column=1, value="CHÚ THÍCH:").font = bold_black
    notes = [
        (GREEN,  "THẮNG kèo"),
        (RED,    "THUA kèo"),
        (YELLOW, "KBQ = Không bình chọn (tính thua)"),
        (BLUE,   "Đã đóng tiền"),
    ]
    for i, (fill, text) in enumerate(notes):
        c = ws.cell(row=note_row + 1 + i, column=1, value=text)
        c.fill   = fill
        c.border = thin
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[note_row + 1 + i].height = 20

    # ---- Lưu file ----
    filename = f"TinhTien_WorldCup2026_{datetime.now(TZ).strftime('%d-%m-%Y')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)

    # Gửi file vào nhóm Telegram
    try:
        with open(filepath, "rb") as f:
            await context.bot.send_document(
                chat_id=GROUP_ID,
                document=f,
                filename=filename,
                caption=f"BẢNG TÍNH TIỀN WORLD CUP 2026\nXuất lúc {datetime.now(TZ).strftime('%d/%m/%Y %H:%M')}\nTổng chưa thu: {total_remain:,}đ"
            )
        await update.message.reply_text(
            f"Đã xuất Excel và gửi vào nhóm!\n"
            f"File cũng được lưu tại: {filepath}"
        )
    except Exception as e:
        await update.message.reply_text(
            f"Đã lưu file tại: {filepath}\n"
            f"Lỗi gửi Telegram: {e}"
        )


def main():
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("themtran",       cmd_newmatch))
    app.add_handler(CommandHandler("xemkeo",      cmd_fetchodds))
    app.add_handler(CommandHandler("laykeo",      cmd_autosetup))
    app.add_handler(CommandHandler("guibinhchon",       cmd_sendpoll))
    app.add_handler(CommandHandler("khoabinhchon",       cmd_lockpoll))
    app.add_handler(CommandHandler("ketqua",         cmd_result))
    app.add_handler(CommandHandler("testapi",        cmd_testapi))
    app.add_handler(CommandHandler("laykequa",    cmd_fetchresult))
    app.add_handler(CommandHandler("dathanhtoan",           cmd_paid))
    app.add_handler(CommandHandler("conno",         cmd_unpaid))
    app.add_handler(CommandHandler("guithele",          cmd_rules))
    app.add_handler(CommandHandler("thele",        cmd_myrules))
    app.add_handler(CommandHandler("thamgia",           cmd_join))
    app.add_handler(CommandHandler("danhsach",        cmd_members))
    app.add_handler(CommandHandler("xoathanhvien",   cmd_removemember))
    app.add_handler(CommandHandler("themthanhvien",      cmd_addmember))
    app.add_handler(CommandHandler("xuatfile",          cmd_excel))
    app.add_handler(CommandHandler("capnhatkeo",     cmd_updateodds))
    app.add_handler(CommandHandler("xoatran",    cmd_deletematch))
    app.add_handler(CommandHandler("suatran",      cmd_editmatch))
    app.add_handler(CommandHandler("suakequa",     cmd_editresult))
    app.add_handler(CommandHandler("suano",       cmd_editdebt))
    app.add_handler(CommandHandler("suadathanhtoan",       cmd_editpaid))
    app.add_handler(CommandHandler("xemtran",      cmd_viewmatch))
    app.add_handler(CommandHandler("danhsachgiai",     cmd_listsports))
    app.add_handler(CommandHandler("xoatrancu",   cmd_clearmatches))
    app.add_handler(CommandHandler("chedo",       cmd_testmode))
    app.add_handler(CommandHandler("resetwc",     cmd_resetforwc))
    app.add_handler(CommandHandler("xoahetdata",      cmd_resetdata))
    app.add_handler(CommandHandler("luudata",        cmd_syncdata))
    app.add_handler(CommandHandler("bangno",      cmd_standings))
    app.add_handler(CommandHandler("lichthidau",        cmd_matches))
    app.add_handler(CommandHandler("lichsu",  cmd_mypredictions))
    app.add_handler(CommandHandler("bieudo",         cmd_bieudo))
    app.add_handler(CommandHandler("thongke",        cmd_thongke))
    app.add_handler(CommandHandler("thongketoi",     cmd_mystat))
    app.add_handler(CommandHandler("huongdan",           cmd_help))
    app.add_handler(PollAnswerHandler(handle_poll_answer))
    logger.info("Bot đang chạy...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
